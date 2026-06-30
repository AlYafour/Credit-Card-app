import csv
import json
from io import StringIO, BytesIO
from django.conf import settings as django_settings
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes, throttle_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.throttling import AnonRateThrottle
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate
from django.core.cache import cache
from django.db.models import Sum, Q, Count
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
from rest_framework.pagination import PageNumberPagination
from .models import Card, Transaction, CashEntry, ChatSession, ChatMessage, WebAuthnCredential, BankPassword, PasswordResetToken, MerchantGroup, Project, AuditLog
from .serializers import (
    UserSerializer, RegisterSerializer, CardSerializer, CardUpdateSerializer,
    TransactionSerializer, CashEntrySerializer, ChatSessionSerializer, ChatMessageSerializer,
    ProjectSerializer, AuditLogSerializer,
)
from .services import encryption_service, parse_card_text, update_card_balance
from .sms_parser import SMSParserEngine


class LoginRateThrottle(AnonRateThrottle):
    rate = '5/minute'


class TransactionPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = 'per_page'
    max_page_size = 200

    def get_paginated_response(self, data):
        return Response({
            'items': data,
            'total': self.page.paginator.count,
            'page': self.page.number,
            'per_page': self.get_page_size(self.request),
            'total_pages': self.page.paginator.num_pages,
            'next': self.get_next_link(),
            'previous': self.get_previous_link(),
        })


def _get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


def audit_log(request, action, model_name, object_id=None, object_repr=None, changes=None):
    """Write an immutable audit entry. Never raises — audit failures must not break the request."""
    try:
        AuditLog.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action=action,
            model_name=model_name,
            object_id=str(object_id) if object_id else None,
            object_repr=object_repr,
            changes=changes,
            ip_address=_get_client_ip(request),
        )
    except Exception:
        pass


# Account lockout constants
MAX_LOGIN_ATTEMPTS = 5
LOCKOUT_DURATION = 300  # 5 minutes in seconds


@api_view(['POST'])
@permission_classes([AllowAny])
def register(request):
    serializer = RegisterSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.save()
        refresh = RefreshToken.for_user(user)
        return Response({
            'access_token': str(refresh.access_token),
            'refresh_token': str(refresh),
            'token_type': 'bearer'
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


def _get_login_data(request):
    """Get email/password from request (JSON body or POST)."""
    data = getattr(request, 'data', None) or {}
    if not isinstance(data, dict):
        data = {}
    if not data and request.body:
        import json
        try:
            data = json.loads(request.body.decode('utf-8')) or {}
        except Exception:
            pass
    email = data.get('email') or (request.POST.get('email') if request.POST else None)
    password = data.get('password') or (request.POST.get('password') if request.POST else None)
    return email, password


@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([LoginRateThrottle])
def login(request):
    try:
        email, password = _get_login_data(request)

        if not email or not password:
            return Response({'detail': 'Email and password required'}, status=status.HTTP_400_BAD_REQUEST)

        # Check for account lockout
        lockout_key = f'login_attempts_{email}'
        attempts = cache.get(lockout_key, 0)
        if attempts >= MAX_LOGIN_ATTEMPTS:
            return Response(
                {'detail': 'Account temporarily locked due to too many failed attempts. Please try again later.'},
                status=status.HTTP_429_TOO_MANY_REQUESTS
            )

        user = authenticate(request, username=email, password=password)
        if user:
            # Reset failed attempts on successful login
            cache.delete(lockout_key)
            refresh = RefreshToken.for_user(user)
            return Response({
                'access_token': str(refresh.access_token),
                'refresh_token': str(refresh),
                'token_type': 'bearer'
            })

        # Increment failed attempts
        cache.set(lockout_key, attempts + 1, LOCKOUT_DURATION)
        return Response({'detail': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)
    except Exception as e:
        import logging
        logging.getLogger(__name__).exception('Login error: %s', e)
        # Never return 500 to client: treat as invalid credentials so UI stays usable
        return Response({'detail': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
@permission_classes([AllowAny])
def refresh_token(request):
    refresh_token = request.data.get('refresh_token')
    if not refresh_token:
        return Response({'detail': 'Refresh token required'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        refresh = RefreshToken(refresh_token)
        access_token = refresh.access_token
        return Response({
            'access_token': str(access_token),
            'refresh_token': str(refresh),
            'token_type': 'bearer'
        })
    except Exception:
        return Response({'detail': 'Invalid refresh token'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def profile(request):
    try:
        if request.method == 'GET':
            serializer = UserSerializer(request.user)
            return Response(serializer.data)
        else:
            serializer = UserSerializer(request.user, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        import logging
        logging.getLogger(__name__).exception('Profile error: %s', e)
        return Response(
            {'detail': 'Invalid or expired token. Please log in again.'},
            status=status.HTTP_401_UNAUTHORIZED
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_password(request):
    current_password = request.data.get('current_password')
    new_password = request.data.get('new_password')

    if not current_password or not new_password:
        return Response({'detail': 'Current and new password required'}, status=status.HTTP_400_BAD_REQUEST)

    if not request.user.check_password(current_password):
        return Response({'detail': 'Current password is incorrect'}, status=status.HTTP_400_BAD_REQUEST)

    if len(new_password) < 8:
        return Response({'detail': 'Password must be at least 8 characters'}, status=status.HTTP_400_BAD_REQUEST)

    from django.contrib.auth.password_validation import validate_password
    from django.core.exceptions import ValidationError as DjangoValidationError
    try:
        validate_password(new_password, request.user)
    except DjangoValidationError as e:
        return Response({'detail': e.messages[0]}, status=status.HTTP_400_BAD_REQUEST)

    request.user.set_password(new_password)
    request.user.save()
    return Response({'message': 'Password changed successfully'})


@api_view(['POST'])
@permission_classes([AllowAny])
def forgot_password_request(request):
    import secrets
    email = (request.data.get('email') or '').strip().lower()
    if not email:
        return Response({'detail': 'Email is required'}, status=status.HTTP_400_BAD_REQUEST)

    from django.contrib.auth import get_user_model
    User = get_user_model()
    try:
        user = User.objects.get(email=email)
    except User.DoesNotExist:
        # Don't reveal whether email exists — always return 200
        return Response({'detail': 'If that email exists, a reset link has been sent.'})

    # Invalidate any existing unused tokens for this user
    PasswordResetToken.objects.filter(user=user, used=False).update(used=True)

    token = secrets.token_urlsafe(32)
    PasswordResetToken.objects.create(
        user=user,
        token=token,
        expires_at=timezone.now() + timedelta(hours=1),
    )

    return Response({
        'detail': 'If that email exists, a reset link has been sent.',
        'token': token,
        'user_name': user.full_name or user.email,
    })


@api_view(['POST'])
@permission_classes([AllowAny])
def reset_password_confirm(request):
    token_str = (request.data.get('token') or '').strip()
    new_password = request.data.get('new_password') or ''

    if not token_str or not new_password:
        return Response({'detail': 'Token and new_password are required'}, status=status.HTTP_400_BAD_REQUEST)

    if len(new_password) < 8:
        return Response({'detail': 'Password must be at least 8 characters'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        reset_token = PasswordResetToken.objects.select_related('user').get(token=token_str)
    except PasswordResetToken.DoesNotExist:
        return Response({'detail': 'Invalid or expired reset link'}, status=status.HTTP_400_BAD_REQUEST)

    if reset_token.used:
        return Response({'detail': 'This reset link has already been used'}, status=status.HTTP_400_BAD_REQUEST)

    if timezone.now() > reset_token.expires_at:
        return Response({'detail': 'Reset link has expired. Please request a new one.'}, status=status.HTTP_400_BAD_REQUEST)

    from django.contrib.auth.password_validation import validate_password
    from django.core.exceptions import ValidationError as DjangoValidationError
    try:
        validate_password(new_password, reset_token.user)
    except DjangoValidationError as e:
        return Response({'detail': e.messages[0]}, status=status.HTTP_400_BAD_REQUEST)

    reset_token.user.set_password(new_password)
    reset_token.user.save()
    reset_token.used = True
    reset_token.save()

    return Response({'message': 'Password reset successfully. You can now log in.'})


class CardViewSet(viewsets.ModelViewSet):
    serializer_class = CardSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    def get_queryset(self):
        return Card.objects.filter(user=self.request.user)
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['reveal'] = self.request.query_params.get('reveal', 'false').lower() == 'true'
        return context
    
    def list(self, request):
        queryset = self.get_queryset()
        bank_name = request.query_params.get('bank_name')
        if bank_name:
            queryset = queryset.filter(bank_name__icontains=bank_name)

        # Support pagination with ?page= or return all with ?all=true
        if request.query_params.get('all', 'true').lower() == 'true':
            serializer = self.get_serializer(queryset, many=True)
            return Response({
                'items': serializer.data,
                'total': queryset.count()
            })

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'items': serializer.data,
            'total': queryset.count()
        })
    
    def retrieve(self, request, pk=None):
        instance = self.get_object()
        reveal = request.query_params.get('reveal', 'false').lower() == 'true'
        if reveal:
            import logging
            logger = logging.getLogger('api.audit')
            logger.info(
                'Card data revealed: user=%s card_id=%s card_last_four=%s ip=%s',
                request.user.email, instance.id, instance.card_last_four,
                request.META.get('REMOTE_ADDR', 'unknown')
            )
        serializer = self.get_serializer(instance, context={'reveal': reveal})
        return Response(serializer.data)
    
    def update(self, request, pk=None, partial=False):
        instance = self.get_object()
        serializer = CardUpdateSerializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(CardSerializer(instance, context={'request': request}).data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def destroy(self, request, pk=None):
        instance = self.get_object()
        instance.is_deleted = True
        instance.save()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=False, methods=['post'])
    def parse_text(self, request):
        text = request.data.get('text', '')
        result = parse_card_text(text)
        return Response(result)

    @action(detail=False, methods=['post'], url_path='scan-image', url_name='scan-image')
    def scan_card_image(self, request):
        """
        Extract card details from a photo using Google Gemini Vision (primary)
        or Anthropic Claude Vision (fallback).
        Image is processed in memory only — never stored.
        """
        import base64
        import logging
        import re as _re
        import time
        import urllib.request
        import urllib.error
        logger = logging.getLogger('api.audit')

        image_data = request.data.get('image')
        if not image_data:
            return Response({'error': 'image is required'}, status=status.HTTP_400_BAD_REQUEST)

        # Strip data URL prefix if present
        media_type = 'image/jpeg'
        if image_data.startswith('data:'):
            try:
                header, image_data = image_data.split(',', 1)
                if 'image/png' in header:
                    media_type = 'image/png'
                elif 'image/webp' in header:
                    media_type = 'image/webp'
                elif 'image/gif' in header:
                    media_type = 'image/gif'
            except ValueError:
                return Response({'error': 'Invalid image data format'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            decoded = base64.b64decode(image_data, validate=True)
        except Exception:
            return Response({'error': 'Invalid base64 encoding'}, status=status.HTTP_400_BAD_REQUEST)

        if len(decoded) > 10 * 1024 * 1024:
            return Response({'error': 'Image too large. Maximum 10MB.'}, status=status.HTTP_400_BAD_REQUEST)

        if not (decoded[:3] == b'\xff\xd8\xff' or decoded[:8] == b'\x89PNG\r\n\x1a\n' or
                decoded[:4] == b'RIFF' or decoded[:6] in (b'GIF87a', b'GIF89a')):
            return Response({'error': 'Invalid image file.'}, status=status.HTTP_400_BAD_REQUEST)

        google_key = getattr(django_settings, 'GOOGLE_API_KEY', '')
        anthropic_key = getattr(django_settings, 'ANTHROPIC_API_KEY', '')

        if not google_key and not anthropic_key:
            return Response(
                {'error': 'Card scanning not configured. Set GOOGLE_API_KEY or ANTHROPIC_API_KEY.'},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )

        logger.info('Card scan attempt: user=%s ip=%s size=%d',
                     request.user.email, request.META.get('REMOTE_ADDR', '?'), len(decoded))

        prompt_text = (
            'Read ALL text visible in this card image. Return a JSON object with these fields '
            '(use null for anything you cannot read):\n'
            '{"card_number":"digits only no spaces","cardholder_name":"name on card",'
            '"expiry_month":"MM","expiry_year":"YY or YYYY",'
            '"cvv":"3-4 digit code","card_network":"visa/mastercard/amex/discover or null",'
            '"bank_name":"issuing bank or null"}\n'
            'Return ONLY the JSON object, nothing else.'
        )

        # Claude needs additional context to understand this is a legitimate use case
        claude_prompt = (
            'You are an OCR assistant inside a personal finance management app called CardVault. '
            'The authenticated user is uploading a photo of THEIR OWN card to store it in their '
            'encrypted personal vault. This is similar to Apple Wallet or Google Pay card scanning. '
            'The user has explicitly consented to this scan.\n\n'
            'Please extract the visible text from this card image and return a JSON object with '
            'these fields (use null for anything not visible):\n'
            '{"card_number":"digits only no spaces","cardholder_name":"name on card",'
            '"expiry_month":"MM","expiry_year":"YY or YYYY",'
            '"cvv":"3-4 digit code","card_network":"visa/mastercard/amex/discover or null",'
            '"bank_name":"issuing bank or null"}\n'
            'Return ONLY the JSON object, nothing else.'
        )

        response_text = None

        # ── Try Anthropic Claude first (more accurate for card OCR) ──
        if anthropic_key:
            try:
                import anthropic
                client = anthropic.Anthropic(api_key=anthropic_key)
                for model_name in ['claude-sonnet-4-6', 'claude-3-5-sonnet-20241022', 'claude-3-5-sonnet-latest']:
                    try:
                        message = client.messages.create(
                            model=model_name, max_tokens=1024,
                            messages=[{'role': 'user', 'content': [
                                {'type': 'image', 'source': {'type': 'base64', 'media_type': media_type, 'data': image_data}},
                                {'type': 'text', 'text': claude_prompt}
                            ]}]
                        )
                        text = message.content[0].text.strip()
                        if text.startswith('{') or '```' in text:
                            response_text = text
                            logger.info('Card scan: Claude %s success user=%s', model_name, request.user.email)
                            break
                        else:
                            logger.warning('Card scan: Claude %s refused, trying next model', model_name)
                            continue
                    except Exception:
                        continue
            except Exception as e:
                logger.warning('Card scan: Anthropic error: %s', str(e))

        # ── Fallback to Google Gemini ─────────────────────────
        if not response_text and google_key:
            gemini_url = (
                'https://generativelanguage.googleapis.com/v1beta/'
                'models/gemini-2.0-flash:generateContent'
                f'?key={google_key}'
            )
            gemini_body = {
                'contents': [{
                    'parts': [
                        {'text': prompt_text},
                        {'inline_data': {'mime_type': media_type, 'data': image_data}}
                    ]
                }]
            }
            payload = json.dumps(gemini_body).encode('utf-8')

            for attempt in range(3):
                try:
                    req = urllib.request.Request(
                        gemini_url, data=payload,
                        headers={'Content-Type': 'application/json'}, method='POST'
                    )
                    with urllib.request.urlopen(req, timeout=30) as resp:
                        resp_body = resp.read().decode('utf-8')
                        data = json.loads(resp_body)
                        candidates = data.get('candidates', [])
                        if candidates:
                            parts = candidates[0].get('content', {}).get('parts', [])
                            if parts:
                                response_text = parts[0].get('text', '').strip()
                                logger.info('Card scan: Gemini success (attempt %d) user=%s', attempt + 1, request.user.email)
                    break
                except urllib.error.HTTPError as e:
                    if e.code == 429 and attempt < 2:
                        wait = (attempt + 1) * 2
                        logger.info('Card scan: Gemini 429, retrying in %ds...', wait)
                        time.sleep(wait)
                        continue
                    logger.warning('Card scan: Gemini HTTP %d (attempt %d): %s', e.code, attempt + 1, str(e))
                    break
                except Exception as e:
                    logger.warning('Card scan: Gemini error (attempt %d): %s', attempt + 1, str(e))
                    break

        if not response_text:
            return Response(
                {'error': 'Could not process image. Please try again.'},
                status=status.HTTP_422_UNPROCESSABLE_ENTITY
            )

        logger.info('Card scan raw (500c): user=%s text=%s', request.user.email, response_text[:500])

        # ── Parse JSON from response ─────────────────────────
        try:
            # Remove markdown fences
            if '```' in response_text:
                fenced = _re.search(r'```(?:json)?\s*\n?(.*?)\n?\s*```', response_text, _re.DOTALL)
                if fenced:
                    response_text = fenced.group(1).strip()

            if not response_text.startswith('{'):
                json_match = _re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', response_text, _re.DOTALL)
                if json_match:
                    response_text = json_match.group(0)

            result = json.loads(response_text)
        except json.JSONDecodeError:
            logger.warning('Card scan: JSON parse failed, raw=%s', response_text[:500])
            return Response(
                {'error': 'Could not extract card details. Please try a clearer photo.'},
                status=status.HTTP_422_UNPROCESSABLE_ENTITY
            )

        # Sanitize output
        allowed = ['card_number', 'cardholder_name', 'expiry_month', 'expiry_year', 'cvv', 'card_network', 'bank_name']
        sanitized = {}
        for f in allowed:
            val = result.get(f)
            if val is not None and str(val).lower() not in ('null', 'none', ''):
                sanitized[f] = str(val).strip()

        if sanitized.get('card_number'):
            sanitized['card_number'] = _re.sub(r'[\s\-]', '', sanitized['card_number'])

        if sanitized.get('card_number') and not sanitized.get('card_network'):
            cn = sanitized['card_number']
            if cn.startswith('4'):
                sanitized['card_network'] = 'visa'
            elif _re.match(r'^(5[1-5]|2[2-7])', cn):
                sanitized['card_network'] = 'mastercard'
            elif _re.match(r'^3[47]', cn):
                sanitized['card_network'] = 'amex'
            elif cn.startswith('6'):
                sanitized['card_network'] = 'discover'

        logger.info('Card scan result: user=%s fields=%s', request.user.email, list(sanitized.keys()))

        if not sanitized:
            return Response(
                {'error': 'Could not read card details. Try a clearer, well-lit photo.'},
                status=status.HTTP_422_UNPROCESSABLE_ENTITY
            )

        return Response(sanitized)

    @action(detail=False, methods=['post'], url_path='extract-document', url_name='extract-document')
    def extract_document(self, request):
        """
        Smart document extractor — accepts any image or PDF (business card, statement,
        benefits brochure, manager photo, etc.) and returns all recognized card-related fields.
        Never stored. Processed in memory only.
        """
        import base64
        import logging
        import urllib.request
        import urllib.error
        import re as _re
        logger = logging.getLogger('api.audit')

        file_data = request.data.get('file')
        file_type = request.data.get('file_type', 'image/jpeg')

        if not file_data:
            return Response({'error': 'file is required'}, status=status.HTTP_400_BAD_REQUEST)

        is_pdf = file_type == 'application/pdf' or (isinstance(file_data, str) and 'application/pdf' in file_data[:50])

        if file_data.startswith('data:'):
            try:
                header, file_data = file_data.split(',', 1)
                if 'pdf' in header:
                    is_pdf = True
                    file_type = 'application/pdf'
                elif 'png' in header:
                    file_type = 'image/png'
                elif 'webp' in header:
                    file_type = 'image/webp'
            except ValueError:
                return Response({'error': 'Invalid file format'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            decoded = base64.b64decode(file_data, validate=True)
        except Exception:
            return Response({'error': 'Invalid base64 encoding'}, status=status.HTTP_400_BAD_REQUEST)

        if len(decoded) > 20 * 1024 * 1024:
            return Response({'error': 'File too large. Maximum 20MB.'}, status=status.HTTP_400_BAD_REQUEST)

        anthropic_key = getattr(django_settings, 'ANTHROPIC_API_KEY', '')
        google_key = getattr(django_settings, 'GOOGLE_API_KEY', '')

        if not anthropic_key and not google_key:
            return Response({'error': 'Document extraction not configured.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        logger.info('Doc extract: user=%s type=%s size=%d', request.user.email, file_type, len(decoded))

        prompt = (
            'You are a smart data extractor for a credit card management app. '
            'Analyze this document carefully — it could be a card benefits brochure, '
            'bank statement, business card, relationship manager photo with text, '
            'or any card-related document.\n\n'
            'Extract ALL useful information and return a JSON object with these fields '
            '(use null for anything not found):\n'
            '{\n'
            '  "card_name": "marketing name of card",\n'
            '  "bank_name": "issuing bank name",\n'
            '  "card_number": "digits only",\n'
            '  "cardholder_name": "name on card",\n'
            '  "expiry_month": "MM",\n'
            '  "expiry_year": "YY",\n'
            '  "card_network": "visa/mastercard/amex/discover",\n'
            '  "credit_limit": "number only",\n'
            '  "annual_fee": "number only",\n'
            '  "late_payment_fee": "number only",\n'
            '  "over_limit_fee": "number only",\n'
            '  "minimum_payment_percentage": "number only e.g. 5",\n'
            '  "statement_date": "day of month 1-31",\n'
            '  "payment_due_date": "day of month 1-31",\n'
            '  "account_manager_name": "full name of relationship manager",\n'
            '  "account_manager_phone": "phone number with country code",\n'
            '  "bank_emails": ["email1", "email2"],\n'
            '  "benefits": [\n'
            '    {"description": "benefit name", "count": "number or null", "notes": "any conditions or notes"}\n'
            '  ]\n'
            '}\n'
            'Be smart: if you see a person\'s name and phone on a business card, '
            'they are likely the account manager. '
            'If you see a list of services, extract them as benefits with count and conditions. '
            'Return ONLY the JSON object, nothing else.'
        )

        response_text = None

        # Try Claude first (better at structured document understanding)
        if anthropic_key:
            try:
                import anthropic
                client = anthropic.Anthropic(api_key=anthropic_key)

                if is_pdf:
                    content = [
                        {'type': 'document', 'source': {'type': 'base64', 'media_type': 'application/pdf', 'data': file_data}},
                        {'type': 'text', 'text': prompt}
                    ]
                else:
                    content = [
                        {'type': 'image', 'source': {'type': 'base64', 'media_type': file_type, 'data': file_data}},
                        {'type': 'text', 'text': prompt}
                    ]

                message = client.messages.create(
                    model='claude-sonnet-4-6', max_tokens=8192,
                    messages=[{'role': 'user', 'content': content}]
                )
                text = message.content[0].text.strip()
                if '{' in text:
                    response_text = text
                    logger.info('Doc extract: Claude success user=%s', request.user.email)
            except Exception as e:
                logger.warning('Doc extract: Claude error: %s', str(e))

        # Fallback to Gemini (images only)
        if not response_text and google_key and not is_pdf:
            try:
                gemini_url = (
                    'https://generativelanguage.googleapis.com/v1beta/'
                    f'models/gemini-2.0-flash:generateContent?key={google_key}'
                )
                gemini_body = {
                    'contents': [{'parts': [
                        {'text': prompt},
                        {'inline_data': {'mime_type': file_type, 'data': file_data}}
                    ]}]
                }
                payload = json.dumps(gemini_body).encode('utf-8')
                req = urllib.request.Request(gemini_url, data=payload,
                    headers={'Content-Type': 'application/json'}, method='POST')
                with urllib.request.urlopen(req, timeout=30) as resp:
                    data = json.loads(resp.read().decode('utf-8'))
                    candidates = data.get('candidates', [])
                    if candidates:
                        parts = candidates[0].get('content', {}).get('parts', [])
                        if parts:
                            response_text = parts[0].get('text', '').strip()
                            logger.info('Doc extract: Gemini success user=%s', request.user.email)
            except Exception as e:
                logger.warning('Doc extract: Gemini error: %s', str(e))

        if not response_text:
            return Response({'error': 'Could not extract information from document.'}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

        # Parse JSON
        try:
            if '```' in response_text:
                fenced = _re.search(r'```(?:json)?\s*\n?(.*?)\n?\s*```', response_text, _re.DOTALL)
                if fenced:
                    response_text = fenced.group(1).strip()
            if not response_text.startswith('{'):
                json_match = _re.search(r'\{.*\}', response_text, _re.DOTALL)
                if json_match:
                    response_text = json_match.group(0)
            result = json.loads(response_text)
        except json.JSONDecodeError:
            return Response({'error': 'Could not parse extracted data.'}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

        logger.info('Doc extract result: user=%s fields=%s', request.user.email, list(result.keys()))
        return Response(result)

    @action(detail=False, methods=['post'], url_path='parse-sms', url_name='parse-sms')
    def parse_sms(self, request):
        """
        Parse SMS message and optionally create transaction
        """
        try:
            sms_body = request.data.get('sms_body', '')
            sender = request.data.get('sender', '')
            received_at = request.data.get('received_at')
            auto_create = request.data.get('auto_create', False)
            card_id = request.data.get('card_id')
            
            if not sms_body:
                return Response({'error': 'sms_body is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Parse SMS
            parser = SMSParserEngine()
            parsed = parser.parse_sms(sms_body, sender)
            
            if not parsed:
                return Response({'error': 'Could not parse SMS message'}, status=status.HTTP_400_BAD_REQUEST)
            
            result = parsed.to_dict()
            
            # Auto-detect card if card_last_four is found and no card_id provided
            matched_card = None
            if parsed.card_last_four and not card_id:
                try:
                    matched_card = Card.objects.filter(
                        user=request.user,
                        card_last_four=parsed.card_last_four,
                    ).first()
                    if matched_card:
                        result['matched_card_id'] = str(matched_card.id)
                        result['matched_card_name'] = matched_card.card_name
                except Exception:
                    pass
            
            # Use matched card or provided card_id
            target_card = matched_card if matched_card else None
            if card_id and not target_card:
                try:
                    target_card = Card.objects.get(id=card_id, user=request.user)
                except Card.DoesNotExist:
                    pass
                except Exception:
                    pass
            
            # Auto-create transaction ONLY if explicitly requested by the client
            should_auto_create = auto_create
            
            if should_auto_create:
                if target_card:
                    try:
                        # Check for duplicate transaction (same amount, date, card, and merchant within 5 minutes)
                        from django.utils import timezone
                        from datetime import timedelta
                        
                        time_window_start = parsed.transaction_date - timedelta(minutes=5)
                        time_window_end = parsed.transaction_date + timedelta(minutes=5)
                        
                        duplicate = Transaction.objects.filter(
                            user=request.user,
                            card=target_card,
                            amount=parsed.amount,
                            transaction_type=parsed.transaction_type,
                            transaction_date__gte=time_window_start,
                            transaction_date__lte=time_window_end,
                        ).first()
                        
                        if duplicate:
                            result['transaction_id'] = str(duplicate.id)
                            result['created'] = False
                            result['duplicate'] = True
                            result['message'] = 'This transaction already exists'
                            result['card_used'] = target_card.card_name
                        else:
                            transaction = Transaction.objects.create(
                                user=request.user,
                                card=target_card,
                                transaction_type=parsed.transaction_type,
                                amount=parsed.amount,
                                currency=parsed.currency,
                                merchant_name=parsed.merchant_name,
                                description=f'Auto-imported from SMS: {sms_body[:100]}',
                                transaction_date=parsed.transaction_date,
                                source='sms_parsed'
                            )
                            # Update card balance
                            if target_card.card_type == 'credit':
                                update_card_balance(target_card)
                            
                            result['transaction_id'] = str(transaction.id)
                            result['created'] = True
                            result['card_used'] = target_card.card_name
                            result['auto_created'] = True
                    except Exception as e:
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.exception("Error creating transaction from SMS")
                        result['error'] = 'Failed to create transaction'
                else:
                    # If auto-create is enabled but no card found, suggest cards
                    if auto_create:
                        result['error'] = 'No card found matching the last 4 digits. Please select a card manually.'
                    else:
                        result['suggested_card_id'] = str(target_card.id) if target_card else None
                        result['suggested_card_name'] = target_card.card_name if target_card else None
            elif target_card:
                result['suggested_card_id'] = str(target_card.id)
                result['suggested_card_name'] = target_card.card_name
            
            return Response(result)
        except Exception:
            import logging
            logging.getLogger(__name__).exception("SMS parsing server error")
            return Response(
                {'error': 'Server error while parsing SMS'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='billing-summary')
    def billing_summary(self, request):
        """Return billing overview for all user's credit cards."""
        cards = Card.objects.filter(user=request.user, is_deleted=False, card_type='credit').order_by('payment_due_date')
        items = []
        total_owed = 0
        total_limit = 0
        total_available = 0
        for card in cards:
            bal = float(card.current_balance) if card.current_balance else 0
            lim = float(card.credit_limit) if card.credit_limit else 0
            available = max(lim - bal, 0)
            total_owed += bal
            total_limit += lim
            total_available += available
            min_pay = None
            if card.minimum_payment:
                min_pay = float(card.minimum_payment)
            elif card.minimum_payment_percentage and card.current_balance:
                min_pay = round(bal * float(card.minimum_payment_percentage) / 100, 2)
            items.append({
                'id': str(card.id), 'card_name': card.card_name,
                'bank_name': card.bank_name, 'card_last_four': card.card_last_four,
                'credit_limit': lim, 'current_balance': bal,
                'available_credit': available,
                'payment_due_date': card.payment_due_date,
                'minimum_payment': min_pay, 'currency': card.balance_currency,
            })
        return Response({'items': items, 'total_owed': total_owed, 'total_credit_limit': total_limit, 'total_available': total_available, 'currency': 'AED'})

    @action(detail=False, methods=['get'], url_path='analytics', url_name='analytics')
    def analytics(self, request):
        """Return comprehensive analytics: spending, payments, points, upcoming dues."""
        from django.db.models import Sum, Count
        from django.db.models.functions import TruncMonth
        from datetime import date, timedelta
        from calendar import monthrange
        import logging
        logger = logging.getLogger('api.audit')

        period = request.query_params.get('period', 'all')
        custom_from = request.query_params.get('from_date')
        custom_to = request.query_params.get('to_date')
        card_id_filter = request.query_params.get('card_id')
        category_filter = request.query_params.get('category')
        today = date.today()

        if custom_from:
            try:
                from datetime import datetime as _dt
                date_from = _dt.strptime(custom_from, '%Y-%m-%d').date()
                period = 'custom'
            except ValueError:
                date_from = None
        elif period == 'month':
            date_from = today.replace(day=1)
        elif period == 'quarter':
            first_month = ((today.month - 1) // 3) * 3 + 1
            date_from = today.replace(month=first_month, day=1)
        elif period == 'year':
            date_from = today.replace(month=1, day=1)
        else:
            date_from = None

        date_to = None
        if custom_to:
            try:
                from datetime import datetime as _dt
                date_to = _dt.strptime(custom_to, '%Y-%m-%d').date()
            except ValueError:
                pass

        txns = Transaction.objects.filter(user=request.user)
        if date_from:
            txns = txns.filter(transaction_date__date__gte=date_from)
        if date_to:
            txns = txns.filter(transaction_date__date__lte=date_to)
        if card_id_filter:
            txns = txns.filter(card_id=card_id_filter)
        if category_filter:
            txns = txns.filter(category__iexact=category_filter)

        purchase_types = ['PURCHASE', 'CASH_WITHDRAWAL', 'CASH_ADVANCE', 'QUASI_CASH']

        total_purchases = txns.filter(transaction_type__in=purchase_types).aggregate(t=Sum('amount'))['t'] or 0
        total_payments  = txns.filter(transaction_type='CARD_PAYMENT').aggregate(t=Sum('amount'))['t'] or 0
        total_refunds   = txns.filter(transaction_type__in=['REFUND', 'REVERSAL', 'CHARGEBACK']).aggregate(t=Sum('amount'))['t'] or 0

        spend_types = purchase_types + ['BANK_FEE', 'FINANCE_CHARGE', 'FOREIGN_EXCHANGE_FEE',
                                         'INSTALLMENT_PRINCIPAL', 'BALANCE_TRANSFER', 'WALLET_TOPUP']

        by_category = list(
            txns.filter(transaction_type__in=spend_types)
            .values('category')
            .annotate(total=Sum('amount'), count=Count('id'))
            .order_by('-total')[:15]
        )

        by_card_qs = list(
            txns.filter(transaction_type__in=spend_types)
            .values('card__id', 'card__card_name', 'card__bank_name',
                    'card__card_last_four', 'card__color_hex',
                    'card__points_earn_rate', 'card__points_value_fils')
            .annotate(total=Sum('amount'), count=Count('id'))
            .order_by('-total')
        )

        # Monthly trend — always last 12 months regardless of period filter
        twelve_ago = today.replace(day=1) - timedelta(days=365)
        monthly_raw = (
            Transaction.objects.filter(user=request.user, transaction_date__date__gte=twelve_ago)
            .annotate(month=TruncMonth('transaction_date'))
            .values('month', 'transaction_type')
            .annotate(total=Sum('amount'))
            .order_by('month')
        )
        monthly_map: dict = {}
        for row in monthly_raw:
            key = row['month'].strftime('%Y-%m')
            if key not in monthly_map:
                monthly_map[key] = {'month': key, 'purchases': 0.0, 'payments': 0.0, 'refunds': 0.0}
            if row['transaction_type'] in spend_types:
                monthly_map[key]['purchases'] += float(row['total'])
            elif row['transaction_type'] == 'CARD_PAYMENT':
                monthly_map[key]['payments'] += float(row['total'])
            elif row['transaction_type'] in ('REFUND', 'REVERSAL', 'CHARGEBACK', 'CASHBACK', 'REWARD_CREDIT'):
                monthly_map[key]['refunds'] += float(row['total'])
        monthly = sorted(monthly_map.values(), key=lambda x: x['month'])

        # Upcoming payment due dates
        credit_cards = Card.objects.filter(user=request.user, card_type__in=['credit', 'covered'])
        upcoming = []
        for card in credit_cards:
            balance = float(card.current_balance or 0)
            if balance <= 0:
                continue
            due_day = card.payment_due_date
            due_date_str = None
            days_until = None
            is_overdue = False
            if due_day:
                try:
                    due_date = today.replace(day=due_day)
                    if due_date <= today:
                        if today.month == 12:
                            due_date = due_date.replace(year=today.year + 1, month=1)
                        else:
                            last_day = monthrange(today.year, today.month + 1)[1]
                            due_date = today.replace(month=today.month + 1, day=min(due_day, last_day))
                    days_until = (due_date - today).days
                    due_date_str = due_date.strftime('%Y-%m-%d')
                    is_overdue = days_until < 0
                except ValueError:
                    pass
            min_pay = 0.0
            if card.minimum_payment:
                min_pay = float(card.minimum_payment)
            elif card.minimum_payment_percentage:
                min_pay = round(balance * float(card.minimum_payment_percentage) / 100, 2)
            upcoming.append({
                'card_id': str(card.id),
                'card_name': card.card_name,
                'bank_name': card.bank_name,
                'card_last_four': card.card_last_four,
                'color_hex': card.color_hex,
                'current_balance': balance,
                'credit_limit': float(card.credit_limit or 0),
                'minimum_payment': min_pay,
                'due_date': due_date_str,
                'days_until': days_until,
                'is_overdue': is_overdue,
                'currency': card.balance_currency or 'AED',
            })
        upcoming.sort(key=lambda x: (x['days_until'] is None, x['days_until'] if x['days_until'] is not None else 999))

        # Points calculation
        total_points = 0.0
        total_points_value = 0.0
        by_card = []
        for r in by_card_qs:
            rate = r.get('card__points_earn_rate') or 1.0
            fils = r.get('card__points_value_fils') or 5.0
            pts = float(r['total']) * rate
            val = pts * fils / 1000
            total_points += pts
            total_points_value += val
            by_card.append({
                'card_id': str(r['card__id']),
                'card_name': r['card__card_name'],
                'bank_name': r['card__bank_name'],
                'last_four': r['card__card_last_four'],
                'color_hex': r['card__color_hex'],
                'total_purchases': float(r['total']),
                'count': r['count'],
                'points_earn_rate': rate,
                'points_value_fils': fils,
                'points_earned': round(pts),
                'points_value_aed': round(val, 2),
            })

        return Response({
            'period': period,
            'date_from': date_from.isoformat() if date_from else None,
            'date_to': date_to.isoformat() if date_to else None,
            'totals': {
                'purchases': float(total_purchases),
                'payments': float(total_payments),
                'refunds': float(total_refunds),
                'net_spending': float(total_purchases) - float(total_refunds),
                'net_after_payments': float(total_purchases) - float(total_refunds) - float(total_payments),
            },
            'by_category': [
                {'category': r['category'] or 'أخرى', 'total': float(r['total']), 'count': r['count']}
                for r in by_category
            ],
            'by_card': by_card,
            'monthly_trend': monthly,
            'upcoming_payments': upcoming,
            'points_summary': {
                'total_earned': round(total_points),
                'total_value_aed': round(total_points_value, 2),
            },
        })

    def _decrypt_pdf(self, pdf_bytes: bytes, password: str) -> bytes:
        """Decrypt a password-protected PDF using pikepdf. Returns decrypted bytes."""
        import io
        try:
            import pikepdf
        except ImportError:
            raise RuntimeError('pikepdf not installed')
        with pikepdf.open(io.BytesIO(pdf_bytes), password=password) as pdf:
            out = io.BytesIO()
            pdf.save(out)
            return out.getvalue()

    @action(detail=False, methods=['post'], url_path='parse-statement', url_name='parse-statement')
    def parse_statement(self, request):
        """Parse a bank statement PDF/image and extract card info + all transactions."""
        import base64
        import re as _re
        import logging
        logger = logging.getLogger('api.audit')

        file_data = request.data.get('file')
        file_type = request.data.get('file_type', 'image/jpeg')
        pdf_password = request.data.get('pdf_password', '')
        save_password = request.data.get('save_password', False)
        bank_name_hint = request.data.get('bank_name_hint', '')

        if not file_data:
            return Response({'error': 'file is required'}, status=status.HTTP_400_BAD_REQUEST)

        is_pdf = file_type == 'application/pdf'
        if file_data.startswith('data:'):
            try:
                header, file_data = file_data.split(',', 1)
                if 'pdf' in header:
                    is_pdf = True
                    file_type = 'application/pdf'
                elif 'png' in header:
                    file_type = 'image/png'
                elif 'webp' in header:
                    file_type = 'image/webp'
            except ValueError:
                return Response({'error': 'Invalid file format'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            decoded = base64.b64decode(file_data, validate=True)
        except Exception:
            return Response({'error': 'Invalid base64 encoding'}, status=status.HTTP_400_BAD_REQUEST)

        if len(decoded) > 25 * 1024 * 1024:
            return Response({'error': 'File too large. Maximum 25MB.'}, status=status.HTTP_400_BAD_REQUEST)

        # Handle PDF password decryption
        if is_pdf:
            pikepdf_available = False
            try:
                import pikepdf as _pikepdf
                import io as _io
                pikepdf_available = True
            except ImportError:
                logger.warning('Statement parse: pikepdf not available, skipping password check')

            if pikepdf_available:
                is_encrypted = False
                try:
                    with _pikepdf.open(_io.BytesIO(decoded)):
                        pass  # opens fine — not encrypted
                except _pikepdf.PasswordError:
                    is_encrypted = True
                except Exception as e:
                    logger.warning('Statement parse: pikepdf open error (non-fatal): %s', str(e))

                if is_encrypted:
                    passwords_to_try = []
                    if pdf_password:
                        passwords_to_try.append(pdf_password)
                    saved_pws = BankPassword.objects.filter(user=request.user)
                    for bp in saved_pws:
                        try:
                            pw = encryption_service.decrypt(bytes(bp.password_encrypted))
                            if pw and pw not in passwords_to_try:
                                passwords_to_try.append(pw)
                        except Exception:
                            pass

                    decrypted = None
                    detected_bank = None
                    for pw in passwords_to_try:
                        try:
                            decrypted_bytes = self._decrypt_pdf(decoded, pw)
                            decrypted = base64.b64encode(decrypted_bytes).decode('utf-8')
                            for bp in saved_pws:
                                try:
                                    if encryption_service.decrypt(bytes(bp.password_encrypted)) == pw:
                                        detected_bank = bp.bank_name
                                        break
                                except Exception:
                                    pass
                            break
                        except Exception:
                            continue

                    if decrypted is None:
                        if not passwords_to_try:
                            return Response(
                                {'error': 'pdf_password_required', 'message': 'هذا الملف محمي بكلمة سر.'},
                                status=status.HTTP_200_OK
                            )
                        return Response(
                            {'error': 'pdf_password_wrong', 'message': 'كلمة السر غير صحيحة.'},
                            status=status.HTTP_200_OK
                        )
                    file_data = decrypted
                    logger.info('Statement parse: PDF decrypted user=%s bank=%s', request.user.email, detected_bank or 'unknown')

        anthropic_key = getattr(django_settings, 'ANTHROPIC_API_KEY', '')
        if not anthropic_key:
            return Response({'error': 'Statement parsing not configured.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        logger.info('Statement parse: user=%s type=%s size=%d', request.user.email, file_type, len(decoded))

        prompt = (
            'You are a bank statement parser for a personal finance app. '
            'Analyze this bank statement carefully and extract ALL information.\n\n'
            'Return a JSON object with exactly this structure:\n'
            '{\n'
            '  "card_info": {\n'
            '    "bank_name": "bank name",\n'
            '    "card_name": "card product name e.g. Platinum Credit Card",\n'
            '    "card_last_four": "last 4 digits of card number only",\n'
            '    "cardholder_name": "name on card/account",\n'
            '    "credit_limit": number_or_null,\n'
            '    "available_balance": number_or_null,\n'
            '    "statement_balance": number_or_null,\n'
            '    "statement_date": day_of_month_1_to_31_or_null,\n'
            '    "payment_due_date": day_of_month_1_to_31_or_null,\n'
            '    "payment_due_full_date": "YYYY-MM-DD or null",\n'
            '    "minimum_payment": number_or_null,\n'
            '    "minimum_payment_percentage": number_or_null,\n'
            '    "annual_fee": number_or_null,\n'
            '    "late_payment_fee": number_or_null,\n'
            '    "over_limit_fee": number_or_null,\n'
            '    "account_manager_name": "name or null",\n'
            '    "account_manager_phone": "phone or null",\n'
            '    "bank_emails": ["email1"],\n'
            '    "currency": "AED",\n'
            '    "statement_period_from": "YYYY-MM-DD or null",\n'
            '    "statement_period_to": "YYYY-MM-DD or null"\n'
            '  },\n'
            '  "transactions": [\n'
            '    {\n'
            '      "date": "YYYY-MM-DD",\n'
            '      "merchant": "merchant name or description",\n'
            '      "amount": number,\n'
            '      "type": "purchase or payment or refund or withdrawal or deposit or transfer",\n'
            '      "currency": "AED",\n'
            '      "category": "Food/Transport/Shopping/etc or null"\n'
            '    }\n'
            '  ]\n'
            '}\n\n'
            'Rules:\n'
            '- Extract EVERY single transaction on the statement\n'
            '- Purchases/spending = type "purchase"\n'
            '- Cash withdrawals = type "withdrawal"\n'
            '- Payments made TO the bank/card account = type "payment"\n'
            '- Refunds/chargebacks credited back = type "refund"\n'
            '- Cash or money deposited INTO the account = type "deposit"\n'
            '- Bank transfers between accounts = type "transfer"\n'
            '- All amounts are positive numbers\n'
            '- Use YYYY-MM-DD for all dates\n'
            '- Return ONLY the JSON object, nothing else'
        )

        try:
            import anthropic as _anthropic
            client = _anthropic.Anthropic(api_key=anthropic_key)
            if is_pdf:
                content = [
                    {'type': 'document', 'source': {'type': 'base64', 'media_type': 'application/pdf', 'data': file_data}},
                    {'type': 'text', 'text': prompt}
                ]
            else:
                content = [
                    {'type': 'image', 'source': {'type': 'base64', 'media_type': file_type, 'data': file_data}},
                    {'type': 'text', 'text': prompt}
                ]
            message = client.messages.create(
                model='claude-sonnet-4-6', max_tokens=8192,
                messages=[{'role': 'user', 'content': content}]
            )
            response_text = message.content[0].text.strip()
            logger.info('Statement parse: Claude success user=%s', request.user.email)
        except Exception as e:
            logger.error('Statement parse: Claude error: %s', str(e))
            return Response({'error': 'Could not parse statement.'}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

        try:
            if '```' in response_text:
                fenced = _re.search(r'```(?:json)?\s*\n?(.*?)\n?\s*```', response_text, _re.DOTALL)
                if fenced:
                    response_text = fenced.group(1).strip()
            if not response_text.startswith('{'):
                json_match = _re.search(r'\{.*\}', response_text, _re.DOTALL)
                if json_match:
                    response_text = json_match.group(0)
            parsed_data = json.loads(response_text)
        except json.JSONDecodeError:
            return Response({'error': 'Could not parse extracted data.'}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

        card_info = parsed_data.get('card_info', {})
        transactions_data = parsed_data.get('transactions', [])

        # Try to match existing card
        matched_card_id = None
        matched_card_name = None
        if card_info.get('card_last_four'):
            existing = Card.objects.filter(
                user=request.user,
                card_last_four=card_info['card_last_four']
            ).first()
            if existing:
                matched_card_id = str(existing.id)
                matched_card_name = existing.card_name

        # Save PDF password if requested
        detected_bank_name = card_info.get('bank_name') or bank_name_hint
        if save_password and pdf_password and detected_bank_name:
            try:
                BankPassword.objects.update_or_create(
                    user=request.user,
                    bank_name=detected_bank_name,
                    defaults={'password_encrypted': encryption_service.encrypt(pdf_password)}
                )
                logger.info('Statement parse: saved password for bank=%s user=%s', detected_bank_name, request.user.email)
            except Exception as e:
                logger.warning('Statement parse: failed to save password: %s', str(e))

        logger.info('Statement parse done: user=%s bank=%s txns=%d matched_card=%s',
                    request.user.email, card_info.get('bank_name'), len(transactions_data), matched_card_id)

        return Response({
            'card_info': card_info,
            'transactions': transactions_data,
            'transaction_count': len(transactions_data),
            'matched_card_id': matched_card_id,
            'matched_card_name': matched_card_name,
            'password_saved': bool(save_password and pdf_password and detected_bank_name),
        })

    @action(detail=False, methods=['post'], url_path='import-statement', url_name='import-statement')
    def import_statement(self, request):
        """Save parsed statement: create/update card + bulk-create transactions."""
        from django.utils import timezone as tz
        from datetime import datetime, timedelta
        import logging
        logger = logging.getLogger('api.audit')

        from .models import Statement as _Statement

        card_info = request.data.get('card_info', {})
        transactions_data = request.data.get('transactions', [])
        card_id = request.data.get('card_id')
        file_data_raw = request.data.get('file')
        file_type_str = request.data.get('file_type', 'application/pdf')
        file_name_str = request.data.get('file_name', 'statement.pdf')

        # Resolve card
        card = None
        if card_id:
            try:
                card = Card.objects.get(id=card_id, user=request.user)
            except Card.DoesNotExist:
                pass

        # Match by last_four (exact)
        if not card and card_info.get('card_last_four'):
            card = Card.objects.filter(
                user=request.user,
                card_last_four=card_info['card_last_four']
            ).first()

        # Fallback: match by bank_name + card_name similarity (handles card renewal / number change)
        if not card and card_info.get('bank_name'):
            bank = card_info['bank_name'].strip()
            cname = (card_info.get('card_name') or '').strip().lower()
            qs = Card.objects.filter(user=request.user, is_deleted=False, bank_name__icontains=bank[:20])
            if cname:
                matched = qs.filter(card_name__icontains=cname[:20]).first()
                if not matched:
                    matched = qs.first()
            else:
                matched = qs.first()
            if matched:
                card = matched
                # Card was renewed — update the last_four to the new number
                if card_info.get('card_last_four') and card.card_last_four != card_info['card_last_four']:
                    from .services import encryption_service as _enc
                    new_last4 = card_info['card_last_four']
                    card.card_last_four = new_last4
                    placeholder = '000000000000' + new_last4
                    card.card_number_encrypted = _enc.encrypt(placeholder)
                    card.save(update_fields=['card_last_four', 'card_number_encrypted'])

        created_card = False
        if not card:
            from .services import encryption_service as _enc
            last_four = card_info.get('card_last_four') or '0000'
            placeholder = '000000000000' + last_four
            card = Card.objects.create(
                user=request.user,
                card_name=card_info.get('card_name') or f"{card_info.get('bank_name', 'Bank')} Card",
                bank_name=card_info.get('bank_name', ''),
                card_type='credit',
                card_number_encrypted=_enc.encrypt(placeholder),
                card_last_four=last_four,
                cardholder_name_encrypted=_enc.encrypt(card_info['cardholder_name']) if card_info.get('cardholder_name') else None,
                balance_currency=card_info.get('currency', 'AED'),
                credit_limit=card_info.get('credit_limit'),
                available_balance=card_info.get('available_balance'),
                statement_date=card_info.get('statement_date'),
                payment_due_date=card_info.get('payment_due_date'),
                minimum_payment=card_info.get('minimum_payment'),
                minimum_payment_percentage=card_info.get('minimum_payment_percentage'),
                annual_fee=card_info.get('annual_fee'),
                late_payment_fee=card_info.get('late_payment_fee'),
                over_limit_fee=card_info.get('over_limit_fee'),
                account_manager_name=card_info.get('account_manager_name'),
                account_manager_phone=card_info.get('account_manager_phone'),
                bank_emails=json.dumps(card_info.get('bank_emails', [])) if card_info.get('bank_emails') else None,
            )
            if card.credit_limit is not None and card.available_balance is not None:
                card.current_balance = float(card.credit_limit) - float(card.available_balance)
                card.save(update_fields=['current_balance'])
            created_card = True
            logger.info('Statement import: created card=%s user=%s', card.id, request.user.email)
        else:
            # Update card fields that are provided
            update_fields = []
            field_map = {
                'credit_limit': 'credit_limit', 'available_balance': 'available_balance',
                'statement_date': 'statement_date', 'payment_due_date': 'payment_due_date',
                'minimum_payment': 'minimum_payment', 'minimum_payment_percentage': 'minimum_payment_percentage',
                'annual_fee': 'annual_fee', 'late_payment_fee': 'late_payment_fee',
                'over_limit_fee': 'over_limit_fee',
            }
            for info_key, model_key in field_map.items():
                val = card_info.get(info_key)
                if val is not None:
                    setattr(card, model_key, val)
                    update_fields.append(model_key)
            if update_fields:
                card.save(update_fields=update_fields + ['updated_at'])

        # Create Statement record
        def _parse_date(s):
            if not s:
                return None
            try:
                from datetime import datetime as _dt
                return _dt.strptime(s[:10], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                return None

        stmt_obj = _Statement.objects.create(
            user=request.user,
            card=card,
            bank_name=card_info.get('bank_name', ''),
            card_name=card_info.get('card_name'),
            card_last_four=card_info.get('card_last_four'),
            cardholder_name=card_info.get('cardholder_name'),
            statement_period_from=_parse_date(card_info.get('statement_period_from')),
            statement_period_to=_parse_date(card_info.get('statement_period_to')),
            statement_balance=card_info.get('statement_balance'),
            available_balance=card_info.get('available_balance'),
            credit_limit=card_info.get('credit_limit'),
            payment_due_full_date=_parse_date(card_info.get('payment_due_full_date')),
            payment_due_day=card_info.get('payment_due_date'),
            minimum_payment=card_info.get('minimum_payment'),
            currency=card_info.get('currency', 'AED'),
        )

        # Save original uploaded file if provided
        if file_data_raw:
            import base64 as _b64
            from io import BytesIO
            from django.core.files.base import ContentFile
            from django.core.files.storage import default_storage
            raw = file_data_raw
            if ',' in raw:
                _, raw = raw.split(',', 1)
            try:
                decoded_file = _b64.b64decode(raw)
                ext_map = {
                    'application/pdf': 'pdf', 'image/jpeg': 'jpg', 'image/jpg': 'jpg',
                    'image/png': 'png', 'image/webp': 'webp',
                }
                ext = ext_map.get(file_type_str, 'pdf')
                storage_path = f'statements/{stmt_obj.id}.{ext}'
                saved_path = default_storage.save(storage_path, ContentFile(decoded_file))
                stmt_obj.file_path = saved_path
                stmt_obj.file_name = file_name_str
                stmt_obj.file_type = file_type_str
                stmt_obj.save(update_fields=['file_path', 'file_name', 'file_type'])
            except Exception as fe:
                logger.warning('Statement import: failed to save file: %s', str(fe))

        # Bulk create transactions
        # Map frontend/AI lowercase type names → canonical uppercase model values
        TYPE_MAP = {
            'purchase': 'PURCHASE',
            'payment': 'CARD_PAYMENT',
            'refund': 'REFUND',
            'withdrawal': 'CASH_WITHDRAWAL',
            'transfer': 'TRANSFER',
            'deposit': 'WALLET_TOPUP',
            'cash_advance': 'CASH_ADVANCE',
            'fee': 'BANK_FEE',
            'interest': 'FINANCE_CHARGE',
            'cashback': 'CASHBACK',
            'reward': 'REWARD_CREDIT',
        }
        VALID_UPPERCASE = set(t[0] for t in Transaction.TRANSACTION_TYPES)
        created_txns = 0
        skipped_txns = 0

        for txn_data in transactions_data:
            try:
                raw_date = txn_data.get('date', '')
                try:
                    txn_date = tz.make_aware(datetime.strptime(raw_date, '%Y-%m-%d'))
                except (ValueError, TypeError):
                    txn_date = tz.now()

                amount = float(txn_data.get('amount', 0))
                if amount <= 0:
                    continue

                raw_type = (txn_data.get('type') or 'purchase').strip().lower()
                txn_type = TYPE_MAP.get(raw_type) or (raw_type.upper() if raw_type.upper() in VALID_UPPERCASE else 'PURCHASE')

                currency = txn_data.get('currency') or (card.balance_currency if card else 'AED')
                merchant = (txn_data.get('merchant') or txn_data.get('description') or '')[:255]

                dup = Transaction.objects.filter(
                    user=request.user, card=card,
                    amount=amount, transaction_type=txn_type,
                    currency=currency,
                    merchant_name=merchant or None,
                    transaction_date__date=txn_date.date(),
                ).first()
                if dup:
                    skipped_txns += 1
                    continue

                Transaction.objects.create(
                    user=request.user, card=card,
                    transaction_type=txn_type, amount=amount,
                    currency=currency, merchant_name=merchant or None,
                    category=txn_data.get('category') or None,
                    transaction_date=txn_date, source='statement_import',
                    statement=stmt_obj,
                )
                created_txns += 1
            except Exception as e:
                logger.warning('Statement import txn error: %s', str(e))
                continue

        # Update statement with final counts
        stmt_obj.transactions_imported = created_txns
        stmt_obj.transactions_skipped = skipped_txns
        stmt_obj.save(update_fields=['transactions_imported', 'transactions_skipped'])

        if card and card.card_type == 'credit':
            update_card_balance(card)

        logger.info('Statement import done: user=%s card=%s created=%d skipped=%d stmt=%s',
                    request.user.email, card.id if card else None, created_txns, skipped_txns, stmt_obj.id)

        from .serializers import CardSerializer as _CS
        return Response({
            'card': _CS(card, context={'request': request}).data if card else None,
            'card_created': created_card,
            'transactions_created': created_txns,
            'transactions_skipped': skipped_txns,
            'total_transactions': len(transactions_data),
            'statement_id': str(stmt_obj.id),
        })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def statements_list(request):
    """List all imported statements with summary, newest first."""
    from .models import Statement as _Statement
    stmts = _Statement.objects.filter(user=request.user).select_related('card').order_by('-imported_at')
    result = []
    for s in stmts:
        result.append({
            'id': str(s.id),
            'bank_name': s.bank_name,
            'card_name': s.card_name,
            'card_last_four': s.card_last_four,
            'cardholder_name': s.cardholder_name,
            'statement_period_from': s.statement_period_from.isoformat() if s.statement_period_from else None,
            'statement_period_to': s.statement_period_to.isoformat() if s.statement_period_to else None,
            'statement_balance': float(s.statement_balance) if s.statement_balance else None,
            'available_balance': float(s.available_balance) if s.available_balance else None,
            'credit_limit': float(s.credit_limit) if s.credit_limit else None,
            'payment_due_full_date': s.payment_due_full_date.isoformat() if s.payment_due_full_date else None,
            'payment_due_day': s.payment_due_day,
            'minimum_payment': float(s.minimum_payment) if s.minimum_payment else None,
            'currency': s.currency,
            'transactions_imported': s.transactions_imported,
            'transactions_skipped': s.transactions_skipped,
            'imported_at': s.imported_at.isoformat(),
            'card_id': str(s.card.id) if s.card else None,
            'card_color': s.card.color_hex if s.card else None,
            'has_file': bool(s.file_path),
            'file_name': s.file_name,
        })
    return Response(result)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def statement_transactions(request, statement_id):
    """List all transactions belonging to a specific statement."""
    from .models import Statement as _Statement
    try:
        stmt = _Statement.objects.get(id=statement_id, user=request.user)
    except _Statement.DoesNotExist:
        from rest_framework.response import Response as _R
        return _R({'error': 'Not found'}, status=404)
    txns = stmt.transactions.filter(is_deleted=False).order_by('-transaction_date')
    result = [{
        'id': str(t.id),
        'date': t.transaction_date.strftime('%Y-%m-%d') if t.transaction_date else None,
        'merchant': t.merchant_name,
        'amount': float(t.amount),
        'type': t.transaction_type,
        'currency': t.currency,
        'category': t.category,
    } for t in txns]
    return Response({'statement_id': statement_id, 'transactions': result, 'count': len(result)})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def statement_file(request, statement_id):
    """Serve the original uploaded file for a statement."""
    from django.http import FileResponse, HttpResponseRedirect, Http404
    from django.core.files.storage import default_storage
    from .models import Statement as _Statement

    try:
        stmt = _Statement.objects.get(id=statement_id, user=request.user)
    except _Statement.DoesNotExist:
        raise Http404

    if not stmt.file_path:
        raise Http404

    # Cloudinary (and any cloud storage): redirect to CDN URL
    if default_storage.__class__.__module__.startswith('cloudinary'):
        if not default_storage.exists(stmt.file_path):
            raise Http404
        return HttpResponseRedirect(default_storage.url(stmt.file_path))

    # Local filesystem fallback
    import os as _os
    abs_path = _os.path.join(django_settings.MEDIA_ROOT, stmt.file_path)
    if not _os.path.exists(abs_path):
        raise Http404
    content_type = stmt.file_type or 'application/octet-stream'
    fh = open(abs_path, 'rb')
    response = FileResponse(fh, content_type=content_type)
    safe_name = (stmt.file_name or 'statement').replace('"', '')
    response['Content-Disposition'] = f'inline; filename="{safe_name}"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def bank_passwords_list(request):
    """List all saved bank passwords for the user (names only, no actual passwords)."""
    items = BankPassword.objects.filter(user=request.user).order_by('bank_name')
    return Response([{'bank_name': bp.bank_name, 'id': str(bp.id), 'updated_at': bp.updated_at} for bp in items])


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bank_passwords_save(request):
    """Save or update a bank password."""
    bank_name = request.data.get('bank_name', '').strip()
    password = request.data.get('password', '').strip()
    if not bank_name or not password:
        return Response({'error': 'bank_name and password are required'}, status=status.HTTP_400_BAD_REQUEST)
    BankPassword.objects.update_or_create(
        user=request.user, bank_name=bank_name,
        defaults={'password_encrypted': encryption_service.encrypt(password)}
    )
    return Response({'status': 'saved', 'bank_name': bank_name})


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def bank_passwords_delete(request, bank_name):
    """Delete a saved bank password."""
    deleted, _ = BankPassword.objects.filter(user=request.user, bank_name=bank_name).delete()
    if deleted:
        return Response({'status': 'deleted'})
    return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)


class TransactionViewSet(viewsets.ModelViewSet):
    serializer_class = TransactionSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = TransactionPagination

    def get_queryset(self):
        include_deleted = self.request.query_params.get('include_deleted') == 'true'
        if include_deleted:
            queryset = Transaction.all_objects.filter(user=self.request.user).select_related('card', 'project')
        else:
            queryset = Transaction.objects.filter(user=self.request.user).select_related('card', 'project')

        card_id = self.request.query_params.get('card_id')
        if card_id:
            queryset = queryset.filter(card_id=card_id)

        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        transaction_type = self.request.query_params.get('transaction_type')

        if start_date:
            queryset = queryset.filter(transaction_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(transaction_date__lte=end_date)
        if transaction_type:
            # Accept both legacy lowercase ('purchase') and canonical uppercase ('PURCHASE')
            _TM = {'purchase': 'PURCHASE', 'payment': 'CARD_PAYMENT', 'refund': 'REFUND',
                   'withdrawal': 'CASH_WITHDRAWAL', 'transfer': 'TRANSFER', 'deposit': 'WALLET_TOPUP',
                   'cash_advance': 'CASH_ADVANCE', 'fee': 'BANK_FEE'}
            queryset = queryset.filter(transaction_type=_TM.get(transaction_type, transaction_type.upper()))

        merchant_name = self.request.query_params.get('merchant_name')
        if merchant_name:
            queryset = queryset.filter(merchant_name__icontains=merchant_name)

        expense_type = self.request.query_params.get('expense_type')
        if expense_type:
            queryset = queryset.filter(expense_type=expense_type)

        merchant_group_id = self.request.query_params.get('merchant_group_id')
        if merchant_group_id:
            queryset = queryset.filter(merchant_group_id=merchant_group_id)

        amount_min = self.request.query_params.get('amount_min')
        amount_max = self.request.query_params.get('amount_max')
        if amount_min:
            queryset = queryset.filter(amount__gte=amount_min)
        if amount_max:
            queryset = queryset.filter(amount__lte=amount_max)

        project_id = self.request.query_params.get('project_id')
        if project_id:
            queryset = queryset.filter(project_id=project_id)

        approval_status = self.request.query_params.get('approval_status')
        if approval_status:
            queryset = queryset.filter(approval_status=approval_status)

        sort = self.request.query_params.get('sort', '-transaction_date')
        valid_sorts = {'transaction_date', '-transaction_date', 'amount', '-amount'}
        if sort not in valid_sorts:
            sort = '-transaction_date'

        return queryset.order_by(sort)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response({'items': serializer.data, 'total': queryset.count()})
    
    def perform_create(self, serializer):
        """Create transaction and update card balance"""
        # Check for duplicate before creating
        validated_data = serializer.validated_data
        card = validated_data.get('card')
        amount = validated_data.get('amount')
        transaction_type = validated_data.get('transaction_type')
        transaction_date = validated_data.get('transaction_date')
        
        if card and amount and transaction_date:
            from django.utils import timezone
            from datetime import timedelta
            
            time_window_start = transaction_date - timedelta(minutes=5)
            time_window_end = transaction_date + timedelta(minutes=5)
            
            duplicate = Transaction.objects.filter(
                user=self.request.user,
                card=card,
                amount=amount,
                transaction_type=transaction_type,
                transaction_date__gte=time_window_start,
                transaction_date__lte=time_window_end,
            ).first()
            
            if duplicate:
                from rest_framework.exceptions import ValidationError
                raise ValidationError({
                    'detail': 'A similar transaction already exists. Please check for duplicates.'
                })
        
        transaction = serializer.save(user=self.request.user)
        
        # Update card balance if transaction is linked to a card
        if transaction.card and transaction.card.card_type == 'credit':
            update_card_balance(transaction.card)
    
    def perform_update(self, serializer):
        """Update transaction and recalculate card balance"""
        old_card = serializer.instance.card
        transaction = serializer.save()
        new_card = transaction.card

        # Recalculate balance for old card if it changed
        if old_card and old_card != new_card:
            update_card_balance(old_card)

        # Update balance for new/current card
        if new_card and new_card.card_type == 'credit':
            update_card_balance(new_card)
    
    def destroy(self, request, pk=None):
        # Get transaction including deleted ones to avoid 404 if already deleted
        try:
            instance = Transaction.all_objects.get(id=pk, user=request.user)
        except Transaction.DoesNotExist:
            return Response(
                {'detail': 'Transaction not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # If already deleted, return success
        if instance.is_deleted:
            return Response(status=status.HTTP_204_NO_CONTENT)
        
        card = instance.card
        instance.is_deleted = True
        instance.save()
        audit_log(request, 'DELETE', 'Transaction', object_id=instance.id, object_repr=str(instance.merchant_name))

        # Recalculate card balance after deletion
        if card:
            update_card_balance(card)

        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'], url_path='submit-approval')
    def submit_approval(self, request, pk=None):
        txn = self.get_object()
        if txn.approval_status not in ('draft', 'rejected'):
            return Response({'detail': 'Already submitted or approved.'}, status=status.HTTP_400_BAD_REQUEST)
        txn.approval_status = 'submitted'
        txn.save(update_fields=['approval_status'])
        audit_log(request, 'SUBMIT', 'Transaction', object_id=txn.id, object_repr=str(txn.merchant_name))
        return Response(TransactionSerializer(txn, context={'request': request}).data)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        txn = self.get_object()
        if txn.approval_status != 'submitted':
            return Response({'detail': 'Transaction must be submitted first.'}, status=status.HTTP_400_BAD_REQUEST)
        txn.approval_status = 'approved'
        txn.approved_by = request.user
        txn.approval_note = request.data.get('note', '')
        txn.save(update_fields=['approval_status', 'approved_by', 'approval_note'])
        audit_log(request, 'APPROVE', 'Transaction', object_id=txn.id, object_repr=str(txn.merchant_name))
        return Response(TransactionSerializer(txn, context={'request': request}).data)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        txn = self.get_object()
        if txn.approval_status != 'submitted':
            return Response({'detail': 'Transaction must be submitted first.'}, status=status.HTTP_400_BAD_REQUEST)
        txn.approval_status = 'rejected'
        txn.approval_note = request.data.get('note', '')
        txn.save(update_fields=['approval_status', 'approval_note'])
        audit_log(request, 'REJECT', 'Transaction', object_id=txn.id, object_repr=str(txn.merchant_name))
        return Response(TransactionSerializer(txn, context={'request': request}).data)

    @action(detail=True, methods=['post'], url_path='upload-receipt')
    def upload_receipt(self, request, pk=None):
        txn = self.get_object()
        receipt = request.FILES.get('receipt')
        if not receipt:
            return Response({'detail': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)
        allowed = {'image/jpeg', 'image/png', 'image/webp', 'application/pdf'}
        if receipt.content_type not in allowed:
            return Response({'detail': 'Only JPEG, PNG, WebP, or PDF files are allowed.'}, status=status.HTTP_400_BAD_REQUEST)
        if receipt.size > 10 * 1024 * 1024:
            return Response({'detail': 'File must be under 10 MB.'}, status=status.HTTP_400_BAD_REQUEST)
        if txn.receipt_file:
            txn.receipt_file.delete(save=False)
        txn.receipt_file = receipt
        txn.save(update_fields=['receipt_file'])
        return Response({'receipt_url': request.build_absolute_uri(txn.receipt_file.url)})

    @action(detail=True, methods=['delete'], url_path='delete-receipt')
    def delete_receipt(self, request, pk=None):
        txn = self.get_object()
        if txn.receipt_file:
            txn.receipt_file.delete(save=False)
            txn.receipt_file = None
            txn.save(update_fields=['receipt_file'])
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'], url_path='restore')
    def restore(self, request, pk=None):
        try:
            txn = Transaction.all_objects.get(id=pk, user=request.user)
        except Transaction.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        if not txn.is_deleted:
            return Response({'detail': 'Transaction is not deleted.'}, status=status.HTTP_400_BAD_REQUEST)
        txn.is_deleted = False
        txn.save(update_fields=['is_deleted'])
        audit_log(request, 'RESTORE', 'Transaction', object_id=txn.id, object_repr=str(txn.merchant_name))
        return Response(TransactionSerializer(txn, context={'request': request}).data)

    @action(detail=False, methods=['get'], url_path='summary/monthly')
    def monthly_summary(self, request):
        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))
        
        start_date = timezone.make_aware(datetime(year, month, 1))
        if month == 12:
            end_date = timezone.make_aware(datetime(year + 1, 1, 1))
        else:
            end_date = timezone.make_aware(datetime(year, month + 1, 1))
        
        transactions = Transaction.objects.filter(
            user=request.user,
            transaction_date__gte=start_date,
            transaction_date__lt=end_date,
        )
        
        expenses = transactions.filter(transaction_type__in=['PURCHASE', 'CASH_WITHDRAWAL', 'CARD_PAYMENT', 'CASH_ADVANCE'])
        income = transactions.filter(transaction_type__in=['REFUND', 'CASHBACK', 'REWARD_CREDIT', 'REVERSAL'])
        
        total_spent = expenses.aggregate(Sum('amount'))['amount__sum'] or 0
        total_income = income.aggregate(Sum('amount'))['amount__sum'] or 0
        
        # Get currency from first transaction or default to AED
        first_transaction = transactions.first()
        currency = first_transaction.currency if first_transaction else 'AED'
        
        return Response({
            'year': year,
            'month': month,
            'total_spent': float(total_spent),
            'total_income': float(total_income),
            'net': float(total_income - total_spent),
            'currency': currency
        })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def merchants_list(request):
    """Return unique merchants from purchase-type transactions only."""
    from django.db.models import Count, Sum, Max
    # Only transaction types that represent an actual merchant interaction
    # Include both legacy lowercase and new uppercase variants
    merchant_types = [
        'purchase', 'refund',                            # legacy
        'PURCHASE', 'REFUND', 'INSTALLMENT_PRINCIPAL',  # new
        'QUASI_CASH', 'PREAUTH_HOLD',
    ]
    # Patterns that indicate a description/charge label, not a real merchant name
    non_merchant_patterns = [
        r'(?i)interest\s+charge',
        r'(?i)finance\s+charge',
        r'(?i)payment\s+by\s+customer',
        r'(?i)minimum\s+payment',
        r'(?i)full\s+payment',
        r'(?i)annual\s+fee',
        r'(?i)late\s+payment\s+fee',
        r'(?i)over\s+limit\s+fee',
        r'\(AED\s+[\d,]+',          # e.g. "(AED 1,900 for 31 days..."
    ]
    queryset = (
        Transaction.objects.filter(
            user=request.user,
            merchant_name__isnull=False,
            transaction_type__in=merchant_types,
        )
        .exclude(merchant_name='')
        .exclude(merchant_name__iexact='purchase')
        .exclude(merchant_name__iexact='refund')
    )
    for pattern in non_merchant_patterns:
        queryset = queryset.exclude(merchant_name__iregex=pattern)

    merchants = (
        queryset
        .values('merchant_name')
        .annotate(
            transaction_count=Count('id'),
            total_amount=Sum('amount'),
            last_transaction_date=Max('transaction_date'),
        )
        .order_by('-transaction_count')
    )
    merchant_list = list(merchants)

    # Attach cached Arabic translations
    from .models import MerchantTranslation
    names = [m['merchant_name'] for m in merchant_list]
    translations = MerchantTranslation.objects.filter(original_name__in=names).values('original_name', 'arabic_name')
    trans_map = {t['original_name']: t['arabic_name'] for t in translations}
    for m in merchant_list:
        m['arabic_name'] = trans_map.get(m['merchant_name'])

    return Response({'items': merchant_list})


def _extract_json_map(text: str) -> dict:
    """Robustly extract a JSON object from AI response text."""
    import re, json as _j
    text = text.strip()
    # Strip markdown code fences (```json ... ``` or ``` ... ```)
    text = re.sub(r'^```[a-z]*\s*', '', text, flags=re.MULTILINE)
    text = re.sub(r'\s*```\s*$', '', text, flags=re.MULTILINE)
    text = text.strip()
    # Try direct parse first
    try:
        return _j.loads(text)
    except Exception:
        pass
    # Find first {...} block
    m = re.search(r'\{[^{}]*\}', text, re.DOTALL)
    if m:
        try:
            return _j.loads(m.group())
        except Exception:
            pass
    return {}


def _call_ai_translate(prompt: str, google_key: str, anthropic_key: str) -> dict:
    """Call Gemini or Claude to translate, return parsed dict."""
    import urllib.request, json as _j, logging
    log = logging.getLogger(__name__)

    if google_key:
        try:
            url = ('https://generativelanguage.googleapis.com/v1beta/'
                   f'models/gemini-2.0-flash:generateContent?key={google_key}')
            body = _j.dumps({'contents': [{'parts': [{'text': prompt}]}],
                             'generationConfig': {'temperature': 0.1, 'maxOutputTokens': 4096}}).encode()
            req = urllib.request.Request(url, data=body, headers={'Content-Type': 'application/json'})
            with urllib.request.urlopen(req, timeout=45) as resp:
                data = _j.loads(resp.read())
            text = data['candidates'][0]['content']['parts'][0]['text']
            result = _extract_json_map(text)
            if result:
                return result
        except Exception as e:
            log.warning(f'Gemini translate failed: {e}')

    if anthropic_key:
        try:
            body = _j.dumps({
                'model': 'claude-sonnet-4-6',
                'max_tokens': 4096,
                'messages': [{'role': 'user', 'content': prompt}],
            }).encode()
            req = urllib.request.Request(
                'https://api.anthropic.com/v1/messages', data=body,
                headers={'Content-Type': 'application/json',
                         'x-api-key': anthropic_key,
                         'anthropic-version': '2023-06-01'})
            with urllib.request.urlopen(req, timeout=45) as resp:
                data = _j.loads(resp.read())
            text = data['content'][0]['text']
            result = _extract_json_map(text)
            if result:
                return result
        except Exception as e:
            log.warning(f'Claude translate failed: {e}')

    return {}


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def translate_merchants(request):
    """Translate merchant names to Arabic using AI and cache in DB."""
    from .models import MerchantTranslation
    import logging
    log = logging.getLogger(__name__)

    names = request.data.get('names', [])
    if not names:
        return Response({'translations': {}})

    # Return already-cached translations
    try:
        existing = MerchantTranslation.objects.filter(
            original_name__in=names).values('original_name', 'arabic_name')
        result = {t['original_name']: t['arabic_name'] for t in existing}
    except Exception as e:
        log.error(f'DB read error in translate_merchants: {e}')
        result = {}

    untranslated = [n for n in names if n not in result]
    if not untranslated:
        return Response({'translations': result})

    google_key = getattr(django_settings, 'GOOGLE_API_KEY', '')
    anthropic_key = getattr(django_settings, 'ANTHROPIC_API_KEY', '')

    # Process in batches of 50 to avoid token limits
    BATCH = 50
    for i in range(0, len(untranslated), BATCH):
        batch = untranslated[i:i + BATCH]
        names_block = '\n'.join(batch)
        prompt = (
            'Translate these merchant/store names to Arabic. '
            'Return ONLY a valid JSON object, no explanation, no markdown.\n'
            'Format: {"ORIGINAL": "عربي", ...}\n'
            'Rules: keep brand names phonetic (e.g. LULU→لولو, IKEA→إيكيا), '
            'add city/country in Arabic if present.\n\n'
            f'{names_block}'
        )
        translated_map = _call_ai_translate(prompt, google_key, anthropic_key)

        # Save to DB
        saved = 0
        for orig, arabic in translated_map.items():
            orig = orig.strip()
            arabic = arabic.strip() if isinstance(arabic, str) else ''
            if orig and arabic and orig in batch:
                try:
                    MerchantTranslation.objects.update_or_create(
                        original_name=orig, defaults={'arabic_name': arabic})
                    result[orig] = arabic
                    saved += 1
                except Exception as e:
                    log.error(f'DB save error for {orig}: {e}')
        log.info(f'translate_merchants: batch {i//BATCH+1}: {saved}/{len(batch)} saved')

    return Response({'translations': result})


class CashEntryViewSet(viewsets.ModelViewSet):
    serializer_class = CashEntrySerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    def get_queryset(self):
        return CashEntry.objects.filter(user=self.request.user, is_deleted=False).order_by('-entry_date')

    def destroy(self, request, pk=None):
        instance = self.get_object()
        instance.is_deleted = True
        instance.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'])
    def balance(self, request):
        entries = CashEntry.objects.filter(user=request.user, is_deleted=False)
        income = entries.filter(entry_type='income').aggregate(Sum('amount'))['amount__sum'] or 0
        expense = entries.filter(entry_type='expense').aggregate(Sum('amount'))['amount__sum'] or 0
        balance = float(income - expense)

        return Response({
            'balance': balance,
            'currency': 'AED'
        })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_data(request):
    """Export user data as CSV or JSON"""
    export_format = request.query_params.get('format', 'json')
    data_type = request.query_params.get('type', 'all')  # all, cards, transactions, cash

    result = {}

    if data_type in ('all', 'cards'):
        cards = Card.objects.filter(user=request.user)
        cards_data = []
        for card in cards:
            cards_data.append({
                'card_name': card.card_name,
                'bank_name': card.bank_name,
                'card_type': card.card_type,
                'card_network': card.card_network or '',
                'card_last_four': card.card_last_four,
                'balance_currency': card.balance_currency,
                'available_balance': str(card.available_balance) if card.available_balance else '',
                'credit_limit': str(card.credit_limit) if card.credit_limit else '',
                'current_balance': str(card.current_balance) if card.current_balance else '',
                'created_at': card.created_at.isoformat(),
            })
        result['cards'] = cards_data

    if data_type in ('all', 'transactions'):
        transactions = Transaction.objects.filter(user=request.user).select_related('card')
        txn_data = []
        for txn in transactions:
            txn_data.append({
                'date': txn.transaction_date.isoformat() if txn.transaction_date else '',
                'type': txn.transaction_type,
                'amount': str(txn.amount),
                'currency': txn.currency,
                'merchant': txn.merchant_name or '',
                'description': txn.description or '',
                'category': txn.category or '',
                'card': txn.card.card_name if txn.card else '',
                'source': txn.source,
            })
        result['transactions'] = txn_data

    if data_type in ('all', 'cash'):
        entries = CashEntry.objects.filter(user=request.user)
        cash_data = []
        for entry in entries:
            cash_data.append({
                'date': entry.entry_date.isoformat() if entry.entry_date else '',
                'type': entry.entry_type,
                'amount': str(entry.amount),
                'currency': entry.currency,
                'description': entry.description or '',
                'category': entry.category or '',
            })
        result['cash_entries'] = cash_data

    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="cardvault_export_{data_type}.csv"'

        writer = csv.writer(response)

        for section_name, section_data in result.items():
            if section_data:
                writer.writerow([f'--- {section_name.upper()} ---'])
                writer.writerow(section_data[0].keys())
                for row in section_data:
                    writer.writerow(row.values())
                writer.writerow([])

        return response
    else:
        response = HttpResponse(
            json.dumps(result, indent=2, ensure_ascii=False),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="cardvault_export_{data_type}.json"'
        return response


class ChatSessionViewSet(viewsets.ModelViewSet):
    serializer_class = ChatSessionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return ChatSession.objects.filter(user=self.request.user).order_by('-created_at')

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class ChatMessageViewSet(viewsets.ModelViewSet):
    serializer_class = ChatMessageSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        session_id = self.request.query_params.get('session_id')
        if session_id:
            return ChatMessage.objects.filter(
                session_id=session_id, session__user=self.request.user
            ).order_by('created_at')
        return ChatMessage.objects.none()
    
    def perform_create(self, serializer):
        session_id = self.request.data.get('session_id')
        try:
            session = ChatSession.objects.get(id=session_id, user=self.request.user)
            serializer.save(session=session)
        except ChatSession.DoesNotExist:
            from rest_framework.exceptions import NotFound
            raise NotFound("Chat session not found")


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def chat_send(request):
    """AI-powered financial chat. Accepts {message, session_id?}, returns {response, session_id}."""
    import time
    import urllib.request
    import urllib.error
    import logging
    logger = logging.getLogger('api.chat')

    user_message = request.data.get('message', '').strip()
    session_id = request.data.get('session_id')
    image_data_url = request.data.get('image')  # optional base64 data URL

    if not user_message:
        return Response({'error': 'Message is required'}, status=status.HTTP_400_BAD_REQUEST)

    # Parse image data URL if provided (format: data:<mime>;base64,<data>)
    image_mime = None
    image_b64 = None
    if image_data_url and isinstance(image_data_url, str) and image_data_url.startswith('data:'):
        try:
            header, image_b64 = image_data_url.split(',', 1)
            image_mime = header.split(':')[1].split(';')[0]  # e.g. "image/png"
        except (ValueError, IndexError):
            image_mime = None
            image_b64 = None

    # If PDF is password-protected, try to decrypt using password found in user message
    if image_mime == 'application/pdf' and image_b64:
        import base64 as _b64_mod, re as _re
        try:
            raw_pdf = _b64_mod.b64decode(image_b64)
            # Extract candidate passwords (sequences of 4+ digits or quoted strings)
            passwords = _re.findall(r'\b\d{4,}\b|["\']([^"\']{4,})["\']', user_message)
            passwords = [p for p in passwords if p]
            import pikepdf
            try:
                pikepdf.open(io.BytesIO(raw_pdf))  # try without password first
            except pikepdf._core.PasswordError:
                decrypted = None
                for pwd in passwords:
                    try:
                        pdf_obj = pikepdf.open(io.BytesIO(raw_pdf), password=pwd)
                        buf = io.BytesIO()
                        pdf_obj.save(buf)
                        decrypted = buf.getvalue()
                        break
                    except Exception:
                        continue
                if decrypted:
                    image_b64 = _b64_mod.b64encode(decrypted).decode('utf-8')
                    logger.info('chat_send: PDF decrypted successfully')
                else:
                    logger.warning('chat_send: could not decrypt PDF — no valid password found in message')
        except Exception as pdf_err:
            logger.warning('chat_send: PDF pre-processing error: %s', pdf_err)

    # Get or create session
    if session_id:
        try:
            session = ChatSession.objects.get(id=session_id, user=request.user)
        except ChatSession.DoesNotExist:
            session = ChatSession.objects.create(user=request.user, title=user_message[:100])
    else:
        session = ChatSession.objects.create(user=request.user, title=user_message[:100])

    # Save user message
    ChatMessage.objects.create(session=session, role='user', content=user_message)

    # Build financial context from user's data
    from django.db.models import Sum, Count, Max, Min
    from django.db.models.functions import TruncMonth

    try:
        user_cards = Card.objects.filter(user=request.user, is_deleted=False)

        # ── Aggregate ALL transactions per card (not just last 50) ──
        all_txns_qs = Transaction.objects.filter(user=request.user, is_deleted=False)
        total_count = all_txns_qs.count()

        # Per-card totals
        card_totals = {}
        for row in all_txns_qs.values('card_id', 'transaction_type').annotate(total=Sum('amount'), cnt=Count('id')):
            cid = str(row['card_id']) if row['card_id'] else '__cash__'
            if cid not in card_totals:
                card_totals[cid] = {'purchases': 0.0, 'payments': 0.0, 'refunds': 0.0, 'withdrawals': 0.0, 'count': 0}
            t = row['transaction_type']
            v = float(row['total'])
            if t in ('PURCHASE', 'CARD_PAYMENT', 'BALANCE_TRANSFER', 'INSTALLMENT_PRINCIPAL', 'BANK_FEE', 'FINANCE_CHARGE', 'FOREIGN_EXCHANGE_FEE', 'QUASI_CASH', 'PREAUTH_HOLD'):
                card_totals[cid]['purchases'] += v
            elif t in ('REFUND', 'CASHBACK', 'REWARD_CREDIT', 'REVERSAL', 'CHARGEBACK', 'ADJUSTMENT', 'PREAUTH_RELEASE'):
                card_totals[cid]['refunds'] += v
            elif t in ('CASH_WITHDRAWAL', 'CASH_ADVANCE'):
                card_totals[cid]['withdrawals'] += v
            elif t in ('WALLET_TOPUP', 'TRANSFER'):
                card_totals[cid]['payments'] += v
            card_totals[cid]['count'] += row['cnt']

        # Per-card latest transaction date
        card_latest = {
            str(r['card_id']): r['latest']
            for r in all_txns_qs.values('card_id').annotate(latest=Max('transaction_date'))
            if r['card_id']
        }

        # Category breakdown (all time, top 15)
        EXPENSE_TYPES = ['PURCHASE', 'CARD_PAYMENT', 'CASH_WITHDRAWAL', 'CASH_ADVANCE',
                         'BALANCE_TRANSFER', 'INSTALLMENT_PRINCIPAL', 'BANK_FEE',
                         'FINANCE_CHARGE', 'FOREIGN_EXCHANGE_FEE', 'QUASI_CASH']
        category_totals = list(
            all_txns_qs.filter(transaction_type__in=EXPENSE_TYPES)
            .values('category').annotate(total=Sum('amount'), cnt=Count('id'))
            .order_by('-total')[:15]
        )

        # Monthly spending (last 6 months)
        from datetime import date as _date, timedelta as _td
        six_ago = _date.today().replace(day=1) - _td(days=180)
        monthly_data = list(
            all_txns_qs.filter(transaction_date__date__gte=six_ago)
            .annotate(month=TruncMonth('transaction_date'))
            .values('month', 'transaction_type')
            .annotate(total=Sum('amount'))
            .order_by('month')
        )
        monthly_map: dict = {}
        for row in monthly_data:
            if not row['month']:
                continue
            key = row['month'].strftime('%Y-%m')
            if key not in monthly_map:
                monthly_map[key] = {'month': key, 'purchases': 0.0, 'payments': 0.0, 'refunds': 0.0}
            t = row['transaction_type']
            if t in ('PURCHASE', 'CARD_PAYMENT', 'CASH_WITHDRAWAL', 'CASH_ADVANCE',
                     'BALANCE_TRANSFER', 'INSTALLMENT_PRINCIPAL', 'BANK_FEE',
                     'FINANCE_CHARGE', 'FOREIGN_EXCHANGE_FEE', 'QUASI_CASH'):
                monthly_map[key]['purchases'] += float(row['total'])
            elif t in ('WALLET_TOPUP', 'TRANSFER'):
                monthly_map[key]['payments'] += float(row['total'])
            elif t in ('REFUND', 'CASHBACK', 'REWARD_CREDIT', 'REVERSAL', 'CHARGEBACK', 'ADJUSTMENT'):
                monthly_map[key]['refunds'] += float(row['total'])

        cards_context = []
        for card in user_cards:
            cid = str(card.id)
            ct = card_totals.get(cid, {})
            net_from_txns = ct.get('purchases', 0) + ct.get('withdrawals', 0) - ct.get('payments', 0) - ct.get('refunds', 0)
            stored_balance = float(card.current_balance) if card.current_balance else None
            latest_txn = card_latest.get(cid)
            card_info = {
                'id': cid,
                'name': card.card_name, 'bank': card.bank_name,
                'type': card.card_type, 'last_four': card.card_last_four,
                'network': card.card_network,
                'credit_limit': float(card.credit_limit) if card.credit_limit else None,
                'stored_balance': stored_balance,
                'computed_balance_from_transactions': round(net_from_txns, 2),
                'available_balance': float(card.available_balance) if card.available_balance else None,
                'currency': card.balance_currency,
                'payment_due_date': card.payment_due_date,
                'statement_date': card.statement_date,
                'minimum_payment': float(card.minimum_payment) if card.minimum_payment else None,
                'min_payment_pct': float(card.minimum_payment_percentage) if card.minimum_payment_percentage else None,
                'txn_summary': {
                    'total_purchases': round(ct.get('purchases', 0), 2),
                    'total_withdrawals': round(ct.get('withdrawals', 0), 2),
                    'total_payments': round(ct.get('payments', 0), 2),
                    'total_refunds': round(ct.get('refunds', 0), 2),
                    'transaction_count': ct.get('count', 0),
                    'latest_transaction': latest_txn.strftime('%Y-%m-%d') if latest_txn else None,
                },
            }
            if card.card_benefits:
                try:
                    card_info['benefits'] = json.loads(card.card_benefits)
                except json.JSONDecodeError:
                    pass
            cards_context.append(card_info)

        # ALL transactions — compact format to stay within context limits
        all_txns_list = all_txns_qs.select_related('card').order_by('-transaction_date')
        txn_context = [
            f"{t.transaction_date.strftime('%Y-%m-%d') if t.transaction_date else '?'}|{t.transaction_type}|{float(t.amount):.2f}|{t.currency}|{t.merchant_name or ''}|{t.card.card_name if t.card else 'Cash'}|{t.card.card_last_four if t.card else ''}|{t.category or ''}"
            for t in all_txns_list
        ]
        # Last 150 transactions with IDs (for delete/update by AI)
        recent_txns_with_ids = [
            {'id': str(t.id), 'date': t.transaction_date.strftime('%Y-%m-%d') if t.transaction_date else '?',
             'type': t.transaction_type, 'amount': float(t.amount), 'currency': t.currency,
             'merchant': t.merchant_name or '', 'card': t.card.card_name if t.card else 'Cash',
             'last4': t.card.card_last_four if t.card else '', 'category': t.category or '',
             'approval': t.approval_status if hasattr(t, 'approval_status') else None}
            for t in list(all_txns_list[:150])
        ]

        # Cash balance
        cash_qs = CashEntry.objects.filter(user=request.user, is_deleted=False)
        cash_in = cash_qs.filter(entry_type='income').aggregate(s=Sum('amount'))['s'] or 0
        cash_out = cash_qs.filter(entry_type='expense').aggregate(s=Sum('amount'))['s'] or 0
        cash_balance = float(cash_in - cash_out)

        # Imported statements (last 20)
        from .models import Statement as _Statement
        stmts_qs = _Statement.objects.filter(user=request.user).select_related('card').order_by('-imported_at')[:20]
        statements_context = [{
            'bank': s.bank_name,
            'card': s.card_name,
            'last_four': s.card_last_four,
            'period': f"{s.statement_period_from} → {s.statement_period_to}" if s.statement_period_from else None,
            'statement_balance': float(s.statement_balance) if s.statement_balance else None,
            'available': float(s.available_balance) if s.available_balance else None,
            'credit_limit': float(s.credit_limit) if s.credit_limit else None,
            'due_date': s.payment_due_full_date.isoformat() if s.payment_due_full_date else None,
            'min_payment': float(s.minimum_payment) if s.minimum_payment else None,
            'txns_imported': s.transactions_imported,
            'imported_at': s.imported_at.strftime('%Y-%m-%d'),
        } for s in stmts_qs]

        # Previous messages for context (last 20)
        prev_msgs = list(ChatMessage.objects.filter(session=session).order_by('-created_at')[:21])
        conversation = [{'role': m.role, 'content': m.content} for m in reversed(prev_msgs) if not (m.role == 'user' and m.content == user_message)]
        if conversation and conversation[-1]['role'] == 'user' and conversation[-1]['content'] == user_message:
            conversation.pop()

        # Build card ID map for action matching
        card_id_map = {}
        for card in user_cards:
            card_id_map[card.card_name.lower()] = str(card.id)
            if card.card_last_four:
                card_id_map[card.card_last_four] = str(card.id)

    except Exception as ctx_err:
        logger.error('chat_send: context build failed: %s', ctx_err, exc_info=True)
        ChatMessage.objects.create(session=session, role='assistant', content='عذراً، حدث خطأ في تحميل بياناتك. حاول مرة أخرى.')
        return Response({'response': 'عذراً، حدث خطأ في تحميل بياناتك. حاول مرة أخرى.', 'session_id': str(session.id), 'actions': []})

    today_str = timezone.now().strftime('%Y-%m-%d')
    system_prompt = f"""أنت CardVault AI — المساعد المالي الذكي لشركة ال يافور للنقليات والمقاولات، أبوظبي، الإمارات العربية المتحدة.
أنت خبير مالي محترف من الدرجة الأولى تشبه مساعد بنك الإمارات دبي الوطني (ENBD) أو بنك أبوظبي الأول (FAB) — دقيق، موثوق، وتتصرف فعلاً.

## بياناتك الكاملة — {len(cards_context)} بطاقة، {total_count} معاملة

### البطاقات وملخصاتها:
{json.dumps(cards_context, ensure_ascii=False, default=str)}

### تحليل الإنفاق:
- أعلى الفئات (كل الوقت): {json.dumps([{{'category': r['category'] or 'Other', 'total': round(float(r['total']),2), 'count': r['cnt']}} for r in category_totals], default=str)}
- الاتجاه الشهري (6 أشهر): {json.dumps(sorted(monthly_map.values(), key=lambda x: x['month']), default=str)}

### الكشوفات المستوردة ({len(statements_context)}):
{json.dumps(statements_context, ensure_ascii=False, default=str)}

### الرصيد النقدي: {cash_balance:.2f} AED

### آخر 150 معاملة بمعرفاتها (للحذف/التعديل):
{json.dumps(recent_txns_with_ids, ensure_ascii=False, default=str)}

### جميع المعاملات ({total_count}) — تنسيق مضغوط (date|type|amount|currency|merchant|card|last4|category):
{chr(10).join(txn_context)}

---

## قواعد الإجابة:
- رُدّ بنفس لغة المستخدم (عربي أو إنجليزي) تلقائياً
- كن دقيقاً بالأرقام — استخدم البيانات الفعلية لا تخمّن
- اعرض المبالغ بوضوح: مثل ١٥٠٠.٠٠ د.إ / 1,500.00 AED
- computed_balance_from_transactions هو الرصيد الأدق (مجموع المعاملات الفعلي)
- stored_balance ما تم إدخاله يدوياً — قد يختلف عن المحسوب
- اليوم: {today_str}
- لا تقل أبداً "لا أستطيع" — لديك صلاحية كاملة للإضافة والتعديل والحذف والتقارير والتصدير

---

## الإجراءات — أنت تتصرف فعلاً!
ضع بلوك الإجراء في نهاية ردّك (بعد النص البشري).
يمكنك تضمين عدة بلوكات في رد واحد.

### إضافة معاملة
عند إضافة أي معاملة أو استخراجها من رسالة SMS أو كشف حساب:
[ACTION:ADD_TRANSACTION]
{{"amount": 150.00, "merchant_name": "Carrefour", "transaction_type": "PURCHASE", "transaction_date": "{today_str}", "card_last_four": "4311", "currency": "AED", "category": "Shopping", "description": "وصف اختياري"}}
[/ACTION]
- transaction_type: PURCHASE | CARD_PAYMENT | CASH_WITHDRAWAL | CASH_ADVANCE | REFUND | REVERSAL | WALLET_TOPUP | TRANSFER | BALANCE_TRANSFER | BANK_FEE | CASHBACK | CHARGEBACK | ADJUSTMENT
- card_last_four: من قائمة البطاقات أعلاه، أو احذفها إن كانت نقداً
- category: Shopping | Food | Transport | Bills | Entertainment | Healthcare | Fuel | Utilities | Government | Salary | Transfer | ATM | Other

### إضافة معاملات دفعة واحدة (استيراد كشف حساب)
عندما يرسل المستخدم كشف حساب بنكي أو قائمة معاملات — استخرج كل المعاملات وأضفها دفعة واحدة:
[ACTION:ADD_TRANSACTIONS_BULK]
{{"card_last_four": "4311", "transactions": [
  {{"amount": 500.00, "merchant_name": "DEWA", "transaction_type": "CARD_PAYMENT", "transaction_date": "{today_str}", "currency": "AED", "category": "Utilities"}},
  {{"amount": 250.00, "merchant_name": "ADNOC", "transaction_type": "PURCHASE", "transaction_date": "{today_str}", "currency": "AED", "category": "Fuel"}}
]}}
[/ACTION]
- أضف جميع المعاملات في بلوك واحد (لا عدة بلوكات منفصلة)
- card_last_four: ينطبق على جميع المعاملات ما لم تحدد غير ذلك

### حذف معاملة
عند حذف معاملة محددة (استخدم الـ id من قائمة "آخر 150 معاملة"):
[ACTION:DELETE_TRANSACTION]
{{"transaction_id": "<uuid من القائمة أعلاه>"}}
[/ACTION]

### تعديل معاملة
عند تعديل أي حقل في معاملة:
[ACTION:UPDATE_TRANSACTION]
{{"transaction_id": "<uuid>", "amount": 200.00, "merchant_name": "اسم جديد", "category": "Food", "transaction_date": "{today_str}", "description": "وصف"}}
[/ACTION]
- أدرج فقط الحقول التي تحتاج تغيير

### إضافة بطاقة
[ACTION:ADD_CARD]
{{"card_name": "Visa Platinum", "bank_name": "Emirates NBD", "card_type": "credit", "card_network": "visa", "card_last_four": "1234", "credit_limit": 50000, "current_balance": 0, "payment_due_date": 25, "balance_currency": "AED"}}
[/ACTION]

### تعديل بطاقة
[ACTION:UPDATE_CARD]
{{"card_id": "<id من قائمة البطاقات>", "current_balance": 5000, "credit_limit": 15000, "payment_due_date": 15}}
[/ACTION]
- أدرج فقط الحقول التي تتغير

### حذف بطاقة
[ACTION:DELETE_CARD]
{{"card_id": "<id>"}}
[/ACTION]

### دمج بطاقتين
عند استبدال بطاقة — انقل جميع المعاملات من القديمة للجديدة ثم احذف القديمة:
[ACTION:MERGE_CARDS]
{{"source_card_id": "<القديمة>", "target_card_id": "<الجديدة>", "delete_source": true}}
[/ACTION]

### مسح كامل للبيانات
يتطلب كلمة تأكيد صريحة (صفّر / امسح كل شيء / clear all / reset data):
[ACTION:CLEAR_ALL_DATA]
{{"clear_transactions": true, "clear_statements": true, "clear_cards": false}}
[/ACTION]

---

## قدرات التقارير والتحليل:
- عند طلب تقرير: اعرض تحليلاً احترافياً مع جداول markdown واضحة
- يمكنك تحليل: الإنفاق الشهري، المقارنة بين البطاقات، أعلى التجار، نسبة استخدام الائتمان، توقعات الدفع القادم
- اقترح دائماً توصيات مالية ذكية بناءً على البيانات الفعلية
- عند وجود كشف حساب: استخرج جميع المعاملات تلقائياً واعرض ملخصاً قبل الإضافة"""

    ai_response = None
    google_key = getattr(django_settings, 'GOOGLE_API_KEY', '')
    anthropic_key = getattr(django_settings, 'ANTHROPIC_API_KEY', '')

    # Try Gemini first
    if google_key:
        gemini_url = (
            'https://generativelanguage.googleapis.com/v1beta/'
            'models/gemini-2.0-flash:generateContent'
            f'?key={google_key}'
        )
        gemini_contents = [
            {'role': 'user', 'parts': [{'text': system_prompt}]},
            {'role': 'model', 'parts': [{'text': 'Understood. I have your financial data ready. How can I help?'}]},
        ]
        for msg in conversation:
            gemini_contents.append({
                'role': 'user' if msg['role'] == 'user' else 'model',
                'parts': [{'text': msg['content']}]
            })
        # Build user message parts (text + optional image)
        user_parts = [{'text': user_message}]
        if image_b64 and image_mime:
            user_parts.append({
                'inline_data': {
                    'mime_type': image_mime,
                    'data': image_b64,
                }
            })
        gemini_contents.append({'role': 'user', 'parts': user_parts})

        payload = json.dumps({
            'contents': gemini_contents,
            'generationConfig': {'temperature': 0.7, 'maxOutputTokens': 8192}
        }).encode('utf-8')

        for attempt in range(3):
            try:
                req = urllib.request.Request(gemini_url, data=payload,
                    headers={'Content-Type': 'application/json'}, method='POST')
                with urllib.request.urlopen(req, timeout=30) as resp:
                    data = json.loads(resp.read().decode('utf-8'))
                    candidates = data.get('candidates', [])
                    if candidates:
                        parts = candidates[0].get('content', {}).get('parts', [])
                        if parts:
                            ai_response = parts[0].get('text', '').strip()
                break
            except urllib.error.HTTPError as e:
                if e.code == 429 and attempt < 2:
                    time.sleep((attempt + 1) * 2)
                    continue
                logger.warning('Chat Gemini HTTP %d', e.code)
                break
            except Exception as e:
                logger.warning('Chat Gemini error: %s', str(e))
                break

    # Fallback to Claude
    if not ai_response and anthropic_key:
        try:
            import anthropic
            client = anthropic.Anthropic(api_key=anthropic_key)
            claude_msgs = [{'role': m['role'], 'content': m['content']} for m in conversation]
            # Build user content (text + optional image)
            if image_b64 and image_mime:
                is_pdf = image_mime == 'application/pdf'
                file_block = {
                    'type': 'document' if is_pdf else 'image',
                    'source': {
                        'type': 'base64',
                        'media_type': image_mime,
                        'data': image_b64,
                    },
                }
                user_content = [
                    {'type': 'text', 'text': user_message},
                    file_block,
                ]
                claude_msgs.append({'role': 'user', 'content': user_content})
            else:
                claude_msgs.append({'role': 'user', 'content': user_message})
            message = client.messages.create(
                model='claude-sonnet-4-6', max_tokens=8192,
                system=system_prompt, messages=claude_msgs,
            )
            ai_response = message.content[0].text.strip()
        except Exception as e:
            logger.warning('Chat Claude error: %s', str(e))

    if not ai_response:
        ai_response = "I'm sorry, I couldn't process your request right now. Please try again."

    # Process actions from AI response
    import re
    from decimal import Decimal
    actions_performed = []
    display_response = ai_response

    # Process ADD_TRANSACTION actions
    if '[ACTION:ADD_TRANSACTION]' in ai_response:
        action_pattern = r'\[ACTION:ADD_TRANSACTION\]\s*(\{.*?\})\s*\[/ACTION\]'
        matches = re.findall(action_pattern, ai_response, re.DOTALL)

        for match in matches:
            try:
                txn_data = json.loads(match)
                amount = txn_data.get('amount')
                if not amount:
                    continue

                # Find the card
                card_obj = None
                card_last_four = txn_data.get('card_last_four', '')
                if card_last_four:
                    card_obj = user_cards.filter(card_last_four=card_last_four).first()

                txn_type = txn_data.get('transaction_type', 'PURCHASE').upper()
                Transaction.objects.create(
                    user=request.user,
                    card=card_obj,
                    amount=Decimal(str(amount)),
                    currency=txn_data.get('currency', 'AED'),
                    transaction_type=txn_type,
                    merchant_name=txn_data.get('merchant_name', ''),
                    category=txn_data.get('category', 'Other'),
                    description=txn_data.get('description', ''),
                    transaction_date=txn_data.get('transaction_date', timezone.now().strftime('%Y-%m-%d')),
                )

                # Update card balance if applicable
                _debit_types = ('PURCHASE', 'CARD_PAYMENT', 'CASH_WITHDRAWAL', 'CASH_ADVANCE',
                                'BALANCE_TRANSFER', 'BANK_FEE', 'FINANCE_CHARGE', 'FOREIGN_EXCHANGE_FEE')
                _credit_types = ('REFUND', 'CASHBACK', 'REVERSAL', 'CHARGEBACK', 'WALLET_TOPUP')
                if card_obj and card_obj.current_balance is not None:
                    if txn_type in _debit_types:
                        card_obj.current_balance = Decimal(str(float(card_obj.current_balance) + amount))
                        card_obj.save(update_fields=['current_balance'])
                    elif txn_type in _credit_types:
                        card_obj.current_balance = Decimal(str(max(0.0, float(card_obj.current_balance) - amount)))
                        card_obj.save(update_fields=['current_balance'])

                actions_performed.append({
                    'type': 'transaction_added',
                    'amount': amount,
                    'merchant': txn_data.get('merchant_name', ''),
                })
            except (json.JSONDecodeError, Exception) as e:
                logger.warning('Chat ADD_TRANSACTION error: %s', str(e))

    # Process ADD_CARD actions
    if '[ACTION:ADD_CARD]' in ai_response:
        action_pattern = r'\[ACTION:ADD_CARD\]\s*(\{.*?\})\s*\[/ACTION\]'
        matches = re.findall(action_pattern, ai_response, re.DOTALL)

        for match in matches:
            try:
                card_data = json.loads(match)
                card_name = card_data.get('card_name', '')
                bank_name = card_data.get('bank_name', '')
                card_last_four = str(card_data.get('card_last_four', '0000'))[-4:].zfill(4)
                card_type = card_data.get('card_type', 'credit')
                if card_type not in ('credit', 'debit', 'prepaid'):
                    card_type = 'credit'

                # Encrypt 16 zeros as placeholder — card_last_four already stored separately
                placeholder_number = '0000000000000000'
                new_card = Card.objects.create(
                    user=request.user,
                    card_name=card_name,
                    bank_name=bank_name,
                    card_type=card_type,
                    card_network=card_data.get('card_network', ''),
                    card_number_encrypted=encryption_service.encrypt(placeholder_number),
                    card_last_four=card_last_four,
                    balance_currency=card_data.get('balance_currency', 'AED'),
                    credit_limit=Decimal(str(card_data['credit_limit'])) if card_data.get('credit_limit') else None,
                    current_balance=Decimal(str(card_data['current_balance'])) if card_data.get('current_balance') is not None else None,
                    payment_due_date=int(card_data['payment_due_date']) if card_data.get('payment_due_date') else None,
                )

                actions_performed.append({
                    'type': 'card_added',
                    'card_name': card_name,
                    'bank_name': bank_name,
                    'card_last_four': card_last_four,
                    'card_id': str(new_card.id),
                })
            except (json.JSONDecodeError, Exception) as e:
                logger.warning('Chat ADD_CARD error: %s', str(e))

    # Process UPDATE_CARD actions
    if '[ACTION:UPDATE_CARD]' in ai_response:
        action_pattern = r'\[ACTION:UPDATE_CARD\]\s*(\{.*?\})\s*\[/ACTION\]'
        matches = re.findall(action_pattern, ai_response, re.DOTALL)
        for match in matches:
            try:
                data = json.loads(match)
                card_id = data.pop('card_id', None)
                if not card_id:
                    continue
                card_obj = Card.objects.filter(id=card_id, user=request.user).first()
                if not card_obj:
                    continue
                allowed_fields = [
                    'card_name', 'bank_name', 'credit_limit', 'current_balance',
                    'available_balance', 'payment_due_date', 'statement_date',
                    'minimum_payment', 'minimum_payment_percentage', 'notes',
                    'color_hex', 'annual_fee', 'late_payment_fee',
                ]
                updated = []
                for field in allowed_fields:
                    if field in data:
                        val = data[field]
                        if field in ('credit_limit', 'current_balance', 'available_balance',
                                     'minimum_payment', 'minimum_payment_percentage',
                                     'annual_fee', 'late_payment_fee') and val is not None:
                            val = Decimal(str(val))
                        setattr(card_obj, field, val)
                        updated.append(field)
                if updated:
                    card_obj.save(update_fields=updated)
                    actions_performed.append({'type': 'card_updated', 'card_id': card_id, 'fields': updated})
            except Exception as e:
                logger.warning('Chat UPDATE_CARD error: %s', str(e))

    # Process DELETE_CARD actions
    if '[ACTION:DELETE_CARD]' in ai_response:
        action_pattern = r'\[ACTION:DELETE_CARD\]\s*(\{.*?\})\s*\[/ACTION\]'
        matches = re.findall(action_pattern, ai_response, re.DOTALL)
        for match in matches:
            try:
                data = json.loads(match)
                card_id = data.get('card_id')
                if not card_id:
                    continue
                card_obj = Card.objects.filter(id=card_id, user=request.user).first()
                if card_obj:
                    card_obj.is_deleted = True
                    card_obj.save(update_fields=['is_deleted'])
                    actions_performed.append({'type': 'card_deleted', 'card_id': card_id, 'card_name': card_obj.card_name})
            except Exception as e:
                logger.warning('Chat DELETE_CARD error: %s', str(e))

    # Process MERGE_CARDS actions
    if '[ACTION:MERGE_CARDS]' in ai_response:
        action_pattern = r'\[ACTION:MERGE_CARDS\]\s*(\{.*?\})\s*\[/ACTION\]'
        matches = re.findall(action_pattern, ai_response, re.DOTALL)
        for match in matches:
            try:
                data = json.loads(match)
                source_id = data.get('source_card_id')
                target_id = data.get('target_card_id')
                if not source_id or not target_id or source_id == target_id:
                    continue
                source = Card.objects.filter(id=source_id, user=request.user).first()
                target = Card.objects.filter(id=target_id, user=request.user).first()
                if not source or not target:
                    continue
                moved = Transaction.objects.filter(card_id=source_id, user=request.user).update(card_id=target_id)
                if data.get('delete_source', True):
                    source.is_deleted = True
                    source.save(update_fields=['is_deleted'])
                actions_performed.append({
                    'type': 'cards_merged',
                    'source_card': source.card_name,
                    'target_card': target.card_name,
                    'transactions_moved': moved,
                })
            except Exception as e:
                logger.warning('Chat MERGE_CARDS error: %s', str(e))

    # Process CLEAR_ALL_DATA actions
    if '[ACTION:CLEAR_ALL_DATA]' in ai_response:
        action_pattern = r'\[ACTION:CLEAR_ALL_DATA\]\s*(\{.*?\})\s*\[/ACTION\]'
        for match in re.finditer(action_pattern, ai_response, re.DOTALL):
            try:
                data = json.loads(match.group(1))
                deleted_txns = deleted_stmts = deleted_cards = 0
                if data.get('clear_transactions', True):
                    deleted_txns, _ = Transaction.objects.filter(user=request.user).delete()
                if data.get('clear_statements', True):
                    from .models import Statement as _Stmt
                    deleted_stmts, _ = _Stmt.objects.filter(user=request.user).delete()
                if data.get('clear_cards', False):
                    deleted_cards = Card.objects.filter(user=request.user).update(is_deleted=True)
                actions_performed.append({
                    'type': 'data_cleared',
                    'transactions_deleted': deleted_txns,
                    'statements_deleted': deleted_stmts,
                    'cards_deleted': deleted_cards,
                })
                logger.info('CLEAR_ALL_DATA: user=%s txns=%d stmts=%d cards=%d',
                            request.user.email, deleted_txns, deleted_stmts, deleted_cards)
            except Exception as e:
                logger.warning('Chat CLEAR_ALL_DATA error: %s', str(e))

    # Process ADD_TRANSACTIONS_BULK actions (statement import)
    if '[ACTION:ADD_TRANSACTIONS_BULK]' in ai_response:
        action_pattern = r'\[ACTION:ADD_TRANSACTIONS_BULK\]\s*(\{.*?\})\s*\[/ACTION\]'
        for match in re.finditer(action_pattern, ai_response, re.DOTALL):
            try:
                data = json.loads(match.group(1))
                card_last_four = data.get('card_last_four', '')
                card_obj = user_cards.filter(card_last_four=card_last_four).first() if card_last_four else None
                txns_list = data.get('transactions', [])
                added = 0
                for td in txns_list:
                    if not td.get('amount'):
                        continue
                    # Per-txn card override
                    per_card_l4 = td.get('card_last_four', card_last_four)
                    per_card = user_cards.filter(card_last_four=per_card_l4).first() if per_card_l4 else card_obj
                    txn_type = td.get('transaction_type', 'PURCHASE').upper()
                    Transaction.objects.create(
                        user=request.user,
                        card=per_card,
                        amount=Decimal(str(td['amount'])),
                        currency=td.get('currency', 'AED'),
                        transaction_type=txn_type,
                        merchant_name=td.get('merchant_name', ''),
                        category=td.get('category', 'Other'),
                        description=td.get('description', ''),
                        transaction_date=td.get('transaction_date', timezone.now().strftime('%Y-%m-%d')),
                    )
                    added += 1
                actions_performed.append({'type': 'bulk_transactions_added', 'count': added, 'card': card_last_four})
                logger.info('ADD_TRANSACTIONS_BULK: user=%s added=%d', request.user.email, added)
            except Exception as e:
                logger.warning('Chat ADD_TRANSACTIONS_BULK error: %s', str(e))

    # Process DELETE_TRANSACTION actions
    if '[ACTION:DELETE_TRANSACTION]' in ai_response:
        action_pattern = r'\[ACTION:DELETE_TRANSACTION\]\s*(\{.*?\})\s*\[/ACTION\]'
        for match in re.finditer(action_pattern, ai_response, re.DOTALL):
            try:
                data = json.loads(match.group(1))
                txn_id = data.get('transaction_id')
                if not txn_id:
                    continue
                txn_obj = Transaction.objects.filter(id=txn_id, user=request.user).first()
                if txn_obj:
                    txn_obj.is_deleted = True
                    txn_obj.save(update_fields=['is_deleted'])
                    actions_performed.append({
                        'type': 'transaction_deleted',
                        'transaction_id': txn_id,
                        'merchant': txn_obj.merchant_name,
                        'amount': float(txn_obj.amount),
                    })
            except Exception as e:
                logger.warning('Chat DELETE_TRANSACTION error: %s', str(e))

    # Process UPDATE_TRANSACTION actions
    if '[ACTION:UPDATE_TRANSACTION]' in ai_response:
        action_pattern = r'\[ACTION:UPDATE_TRANSACTION\]\s*(\{.*?\})\s*\[/ACTION\]'
        for match in re.finditer(action_pattern, ai_response, re.DOTALL):
            try:
                data = json.loads(match.group(1))
                txn_id = data.pop('transaction_id', None)
                if not txn_id:
                    continue
                txn_obj = Transaction.objects.filter(id=txn_id, user=request.user).first()
                if not txn_obj:
                    continue
                allowed = ['amount', 'merchant_name', 'category', 'transaction_date', 'description',
                           'transaction_type', 'currency', 'expense_type']
                updated = []
                for field in allowed:
                    if field in data:
                        val = data[field]
                        if field == 'amount':
                            val = Decimal(str(val))
                        if field == 'transaction_type':
                            val = str(val).upper()
                        setattr(txn_obj, field, val)
                        updated.append(field)
                if updated:
                    txn_obj.save(update_fields=updated)
                    actions_performed.append({'type': 'transaction_updated', 'transaction_id': txn_id, 'fields': updated})
            except Exception as e:
                logger.warning('Chat UPDATE_TRANSACTION error: %s', str(e))

    # Strip all action blocks from the display response
    display_response = re.sub(r'\[ACTION:[A-Z_]+\].*?\[/ACTION\]', '', ai_response, flags=re.DOTALL).strip()

    # Save AI response (clean version without action blocks)
    ChatMessage.objects.create(session=session, role='assistant', content=display_response)

    response_data = {'response': display_response, 'session_id': str(session.id)}
    if actions_performed:
        response_data['actions'] = actions_performed
    return Response(response_data)


# ---------------------------------------------------------------------------
# WebAuthn / FIDO2 Biometric Authentication
# ---------------------------------------------------------------------------

import os
import base64
import hashlib
import struct


def _b64url_encode(data: bytes) -> str:
    """Base64url-encode bytes without padding."""
    return base64.urlsafe_b64encode(data).rstrip(b'=').decode('ascii')


def _b64url_decode(s: str) -> bytes:
    """Base64url-decode a string, adding padding as needed."""
    s += '=' * (4 - len(s) % 4)
    return base64.urlsafe_b64decode(s)


def _get_rp_id(request):
    """Get the WebAuthn Relying Party ID.
    Priority:
    1. WEBAUTHN_RP_ID env var (if not 'localhost')
    2. rp_id query param or POST body field (sent by frontend)
    3. HTTP_ORIGIN header
    4. HTTP_REFERER header
    5. Falls back to configured value ('localhost')
    """
    from urllib.parse import urlparse

    configured = getattr(django_settings, 'WEBAUTHN_RP_ID', 'localhost')
    if configured and configured != 'localhost':
        return configured

    # Check query params (GET requests) or POST body
    rp_id_param = (
        request.query_params.get('rp_id')
        or (request.data.get('rp_id') if hasattr(request, 'data') else None)
    )
    if rp_id_param and rp_id_param != 'localhost':
        return rp_id_param

    # Auto-detect from Origin/Referer header
    for header in ('HTTP_ORIGIN', 'HTTP_REFERER'):
        header_val = request.META.get(header, '')
        if header_val:
            parsed = urlparse(header_val)
            if parsed.hostname and parsed.hostname != 'localhost':
                return parsed.hostname

    return configured


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def webauthn_register_options(request):
    """
    Generate WebAuthn registration options (challenge + relying party info).
    The client uses these to call navigator.credentials.create().
    """
    user = request.user

    # Generate a cryptographically random challenge
    challenge = os.urandom(32)
    cache_key = f'webauthn_reg_{user.id}'
    cache.set(cache_key, challenge, 300)  # 5 minutes

    # Relying party configuration
    rp_id = _get_rp_id(request)
    rp_name = 'CardVault'

    # Build list of credentials to exclude (user already registered these)
    existing_credentials = WebAuthnCredential.objects.filter(user=user)
    exclude_credentials = [
        {
            'type': 'public-key',
            'id': cred.credential_id,
        }
        for cred in existing_credentials
    ]

    display_name = user.full_name or user.email

    return Response({
        'challenge': _b64url_encode(challenge),
        'rp': {
            'id': rp_id,
            'name': rp_name,
        },
        'user': {
            'id': _b64url_encode(str(user.id).encode('utf-8')),
            'name': user.email,
            'displayName': display_name,
        },
        'pubKeyCredParams': [
            {'type': 'public-key', 'alg': -7},    # ES256
            {'type': 'public-key', 'alg': -257},   # RS256
        ],
        'authenticatorSelection': {
            'authenticatorAttachment': 'platform',
            'userVerification': 'required',
        },
        'timeout': 60000,
        'excludeCredentials': exclude_credentials,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def webauthn_register_verify(request):
    """
    Verify and store a new WebAuthn credential after the browser's
    navigator.credentials.create() call succeeds.

    NOTE: This uses a "trust the client" approach -- the browser's WebAuthn
    API has already verified the attestation locally. Full server-side CBOR/
    COSE attestation verification would require the ``fido2`` library.
    """
    user = request.user

    credential_id = request.data.get('credential_id')
    public_key = request.data.get('public_key')
    sign_count = request.data.get('sign_count', 0)
    device_name = request.data.get('device_name', '')

    if not credential_id or not public_key:
        return Response(
            {'detail': 'credential_id and public_key are required'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Verify that a registration challenge was issued for this user
    cache_key = f'webauthn_reg_{user.id}'
    challenge = cache.get(cache_key)
    if challenge is None:
        return Response(
            {'detail': 'Registration challenge expired or not found. Please restart registration.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Consume the challenge so it cannot be reused
    cache.delete(cache_key)

    # Check for duplicate credential
    if WebAuthnCredential.objects.filter(credential_id=credential_id).exists():
        return Response(
            {'detail': 'This credential is already registered.'},
            status=status.HTTP_409_CONFLICT,
        )

    # Store the credential
    credential = WebAuthnCredential.objects.create(
        user=user,
        credential_id=credential_id,
        public_key=public_key,
        sign_count=sign_count,
        device_name=device_name or None,
    )

    return Response({
        'message': 'Credential registered successfully',
        'credential_id': str(credential.id),
        'device_name': credential.device_name,
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([AllowAny])
def webauthn_login_options(request):
    """
    Generate WebAuthn login (assertion) options for a given email.
    The client uses these to call navigator.credentials.get().
    """
    email = request.data.get('email')
    if not email:
        return Response(
            {'detail': 'Email is required'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Import User model
    from .models import User

    try:
        user = User.objects.get(email=email)
    except User.DoesNotExist:
        # Return a generic error to avoid user enumeration
        return Response(
            {'detail': 'No WebAuthn credentials found for this account.'},
            status=status.HTTP_404_NOT_FOUND,
        )

    credentials = WebAuthnCredential.objects.filter(user=user)
    if not credentials.exists():
        return Response(
            {'detail': 'No WebAuthn credentials found for this account.'},
            status=status.HTTP_404_NOT_FOUND,
        )

    # Generate challenge
    challenge = os.urandom(32)
    cache_key = f'webauthn_login_{email}'
    cache.set(cache_key, challenge, 300)  # 5 minutes

    rp_id = _get_rp_id(request)

    allow_credentials = [
        {
            'type': 'public-key',
            'id': cred.credential_id,
        }
        for cred in credentials
    ]

    return Response({
        'challenge': _b64url_encode(challenge),
        'rpId': rp_id,
        'allowCredentials': allow_credentials,
        'timeout': 60000,
        'userVerification': 'required',
    })


@api_view(['POST'])
@permission_classes([AllowAny])
def webauthn_login_verify(request):
    """
    Verify a WebAuthn login assertion and issue JWT tokens.

    NOTE: This uses a simplified verification approach -- it confirms the
    credential exists, belongs to the correct user, and enforces sign_count
    monotonicity for replay protection. Full CBOR/COSE signature verification
    would require the ``fido2`` library.
    """
    email = request.data.get('email')
    credential_id = request.data.get('credential_id')
    authenticator_data = request.data.get('authenticator_data')
    client_data_json = request.data.get('client_data_json')
    signature = request.data.get('signature')

    if not all([email, credential_id, authenticator_data, client_data_json, signature]):
        return Response(
            {'detail': 'email, credential_id, authenticator_data, client_data_json, and signature are required'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Verify that a login challenge was issued for this email
    cache_key = f'webauthn_login_{email}'
    challenge = cache.get(cache_key)
    if challenge is None:
        return Response(
            {'detail': 'Login challenge expired or not found. Please restart authentication.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Consume the challenge
    cache.delete(cache_key)

    # Look up the credential
    try:
        credential = WebAuthnCredential.objects.select_related('user').get(
            credential_id=credential_id,
        )
    except WebAuthnCredential.DoesNotExist:
        return Response(
            {'detail': 'Credential not found.'},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    # Verify the credential belongs to the correct user
    if credential.user.email != email:
        return Response(
            {'detail': 'Credential does not match the provided email.'},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    # Verify the user account is active
    if not credential.user.is_active:
        return Response(
            {'detail': 'User account is disabled.'},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    # Extract and verify sign_count from authenticator_data for replay protection
    # authenticator_data is base64url-encoded; the sign count is a 4-byte
    # big-endian integer at offset 33 (after 32-byte rpIdHash + 1-byte flags).
    try:
        auth_data_bytes = _b64url_decode(authenticator_data)
        if len(auth_data_bytes) >= 37:
            new_sign_count = struct.unpack('>I', auth_data_bytes[33:37])[0]
        else:
            new_sign_count = 0
    except Exception:
        new_sign_count = 0

    # Replay protection: sign_count must be strictly greater than stored value,
    # unless both are 0 (some authenticators don't implement counters).
    if credential.sign_count > 0 and new_sign_count <= credential.sign_count:
        return Response(
            {'detail': 'Potential credential cloning detected (sign count mismatch).'},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    # Update credential metadata
    credential.sign_count = new_sign_count
    credential.last_used_at = timezone.now()
    credential.save(update_fields=['sign_count', 'last_used_at'])

    # Issue JWT tokens
    refresh = RefreshToken.for_user(credential.user)

    return Response({
        'access_token': str(refresh.access_token),
        'refresh_token': str(refresh),
        'token_type': 'bearer',
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def webauthn_list_credentials(request):
    """List all registered biometric credentials for the current user."""
    credentials = WebAuthnCredential.objects.filter(user=request.user).order_by('-created_at')
    data = [
        {
            'id': str(cred.id),
            'device_name': cred.device_name or 'Unknown Device',
            'created_at': cred.created_at.isoformat() if cred.created_at else None,
            'last_used_at': cred.last_used_at.isoformat() if cred.last_used_at else None,
        }
        for cred in credentials
    ]
    return Response(data)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def webauthn_delete_credential(request, pk):
    """Delete a specific biometric credential belonging to the current user."""
    try:
        credential = WebAuthnCredential.objects.get(id=pk, user=request.user)
        credential.delete()
        return Response({'detail': 'Credential deleted.'}, status=status.HTTP_200_OK)
    except WebAuthnCredential.DoesNotExist:
        return Response({'detail': 'Credential not found.'}, status=status.HTTP_404_NOT_FOUND)


# ─────────────────────────────────────────────────────────────
# OpenAI Realtime – ephemeral session token endpoint
# ─────────────────────────────────────────────────────────────

REALTIME_MODEL = 'gpt-4o-realtime-preview-2024-12-17'


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def realtime_session(request):
    """
    Create an ephemeral OpenAI Realtime session token injected with the user's full
    financial context (same data as chat_send) + tools for adding transactions/cards.
    The API key stays on the server; the browser only receives the short-lived client_secret.
    """
    import urllib.request
    import urllib.error

    openai_key = getattr(django_settings, 'OPENAI_API_KEY', '')
    if not openai_key:
        return Response(
            {'error': 'OpenAI Realtime not configured. Add OPENAI_API_KEY to environment.'},
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    # ── Build financial context (identical to chat_send) ──────────────────────
    user_cards = Card.objects.filter(user=request.user, is_deleted=False)
    cards_context = []
    card_id_map = {}   # last_four / name → card id (for tool execution)
    for card in user_cards:
        card_id_map[card.card_name.lower()] = str(card.id)
        if card.card_last_four:
            card_id_map[card.card_last_four] = str(card.id)
        card_info = {
            'id': str(card.id),
            'name': card.card_name,
            'bank': card.bank_name,
            'type': card.card_type,
            'last_four': card.card_last_four,
            'network': card.card_network,
            'credit_limit': float(card.credit_limit) if card.credit_limit else None,
            'current_balance': float(card.current_balance) if card.current_balance else None,
            'available_balance': float(card.available_balance) if card.available_balance else None,
            'currency': card.balance_currency,
            'payment_due_date': card.payment_due_date,
            'minimum_payment': float(card.minimum_payment) if card.minimum_payment else None,
        }
        if card.card_benefits:
            try:
                card_info['benefits'] = json.loads(card.card_benefits)
            except json.JSONDecodeError:
                pass
        cards_context.append(card_info)

    recent_txns = Transaction.objects.filter(
        user=request.user, is_deleted=False
    ).select_related('card').order_by('-transaction_date')[:30]
    txn_context = [{
        'type': t.transaction_type, 'amount': float(t.amount),
        'currency': t.currency, 'merchant': t.merchant_name,
        'date': t.transaction_date.strftime('%Y-%m-%d') if t.transaction_date else None,
        'card': t.card.card_name if t.card else 'Cash',
        'category': t.category,
    } for t in recent_txns]

    cash_qs = CashEntry.objects.filter(user=request.user, is_deleted=False)
    cash_in  = cash_qs.filter(entry_type='income').aggregate(s=Sum('amount'))['s'] or 0
    cash_out = cash_qs.filter(entry_type='expense').aggregate(s=Sum('amount'))['s'] or 0
    cash_balance = float(cash_in - cash_out)

    today = timezone.now().strftime('%Y-%m-%d')

    instructions = f"""You are CardVault AI, a smart voice financial assistant.

## User's Cards ({len(cards_context)} cards):
{json.dumps(cards_context, ensure_ascii=False, default=str)}

## Recent Transactions (last {len(txn_context)}):
{json.dumps(txn_context, ensure_ascii=False, default=str)}

## Cash Balance: {cash_balance} AED

## Rules:
- Always respond in the SAME language the user speaks — Arabic if Arabic, English if English.
- Be concise and clear — this is a voice conversation.
- Spell out numbers clearly. Format amounts: e.g. "ألف وخمسمائة درهم" or "1,500 AED".
- Today is {today}.
- Never make up data — only use the financial data provided above.
- You CAN add transactions and cards — use the provided functions when asked.
- When you add something, confirm it verbally to the user."""

    # ── Realtime Tools (Function Calling) ─────────────────────────────────────
    tools = [
        {
            'type': 'function',
            'name': 'add_transaction',
            'description': (
                'Add a new financial transaction for the user. '
                'Call this when the user says they spent money, received money, '
                'or wants to record any purchase, payment, or transfer.'
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'amount': {
                        'type': 'number',
                        'description': 'Transaction amount (positive number)',
                    },
                    'merchant_name': {
                        'type': 'string',
                        'description': 'Merchant or transaction description',
                    },
                    'transaction_type': {
                        'type': 'string',
                        'enum': ['purchase', 'withdrawal', 'payment', 'refund', 'transfer', 'deposit'],
                        'description': 'Type of transaction',
                    },
                    'card_last_four': {
                        'type': 'string',
                        'description': 'Last 4 digits of the card used, if applicable',
                    },
                    'currency': {
                        'type': 'string',
                        'description': 'Currency code (default: AED)',
                    },
                    'category': {
                        'type': 'string',
                        'enum': ['Shopping', 'Food', 'Transport', 'Bills', 'Entertainment', 'Transfer', 'ATM', 'Other'],
                    },
                    'transaction_date': {
                        'type': 'string',
                        'description': f'Date YYYY-MM-DD (default: today {today})',
                    },
                },
                'required': ['amount', 'merchant_name', 'transaction_type'],
            },
        },
        {
            'type': 'function',
            'name': 'add_card',
            'description': (
                'Add a new credit or debit card for the user. '
                'Call this when the user asks to save or add a new card.'
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'card_name': {
                        'type': 'string',
                        'description': 'Descriptive card name e.g. "Visa Platinum"',
                    },
                    'bank_name': {
                        'type': 'string',
                        'description': 'Issuing bank name',
                    },
                    'card_type': {
                        'type': 'string',
                        'enum': ['credit', 'debit', 'prepaid'],
                    },
                    'card_network': {
                        'type': 'string',
                        'enum': ['visa', 'mastercard', 'amex', 'other'],
                    },
                    'card_last_four': {
                        'type': 'string',
                        'description': 'Last 4 digits of the card',
                    },
                    'credit_limit': {
                        'type': 'number',
                        'description': 'Credit limit (credit cards only)',
                    },
                    'currency': {
                        'type': 'string',
                        'description': 'Card currency (default: AED)',
                    },
                    'payment_due_date': {
                        'type': 'integer',
                        'description': 'Day of month for payment due date (1-31)',
                    },
                },
                'required': ['card_name', 'bank_name', 'card_type'],
            },
        },
    ]

    payload = {
        'model': REALTIME_MODEL,
        'modalities': ['audio', 'text'],
        'voice': 'verse',
        'instructions': instructions,
        'tools': tools,
        'tool_choice': 'auto',
        'max_response_output_tokens': 512,
        'input_audio_transcription': {
            'model': 'gpt-4o-mini-transcribe',
        },
        'turn_detection': {
            'type': 'server_vad',
            'silence_duration_ms': 600,
            'prefix_padding_ms': 200,
            'threshold': 0.5,
        },
    }

    try:
        req = urllib.request.Request(
            'https://api.openai.com/v1/realtime/sessions',
            data=json.dumps(payload, default=str).encode('utf-8'),
            headers={
                'Authorization': f'Bearer {openai_key}',
                'Content-Type': 'application/json',
            },
            method='POST',
        )
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read().decode('utf-8'))

        return Response({
            'client_secret': data['client_secret']['value'],
            'expires_at': data['client_secret'].get('expires_at'),
            'model': REALTIME_MODEL,
            # Pass card_id_map to frontend so it can resolve card names → IDs for tool calls
            'card_id_map': card_id_map,
        })

    except urllib.error.HTTPError as e:
        detail = e.read().decode('utf-8', errors='replace')
        return Response(
            {'error': f'OpenAI error {e.code}', 'detail': detail},
            status=status.HTTP_502_BAD_GATEWAY,
        )
    except Exception as exc:
        return Response({'error': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ─── Excel Export / Import ───────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def transactions_export_excel(request):
    """Export all (filtered) transactions as .xlsx"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from .models import Transaction as TxnModel

    qs = TxnModel.objects.filter(user=request.user, is_deleted=False).select_related('card', 'merchant_group').order_by('-transaction_date')

    # Apply same filters as list view
    card_id = request.GET.get('card_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    txn_type = request.GET.get('transaction_type')
    merchant = request.GET.get('merchant_name')
    expense_type = request.GET.get('expense_type')
    merchant_group_id = request.GET.get('merchant_group_id')

    if card_id:
        qs = qs.filter(card_id=card_id)
    if start_date:
        qs = qs.filter(transaction_date__date__gte=start_date)
    if end_date:
        qs = qs.filter(transaction_date__date__lte=end_date)
    if txn_type:
        qs = qs.filter(transaction_type=txn_type)
    if merchant:
        qs = qs.filter(merchant_name__icontains=merchant)
    if expense_type:
        qs = qs.filter(expense_type=expense_type)
    if merchant_group_id:
        qs = qs.filter(merchant_group_id=merchant_group_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions'

    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    center = Alignment(horizontal='center', vertical='center')

    headers = ['Date', 'Type', 'Merchant', 'Description', 'Category', 'Amount', 'Currency', 'Card', 'Expense Type', 'Basket', 'Source']
    col_widths = [20, 22, 28, 30, 18, 14, 10, 24, 16, 20, 12]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[chr(64 + i)].width = w

    for row_num, txn in enumerate(qs, 2):
        ws.append([
            txn.transaction_date.strftime('%Y-%m-%d %H:%M') if txn.transaction_date else '',
            txn.transaction_type,
            txn.merchant_name or '',
            txn.description or '',
            txn.category or '',
            float(txn.amount),
            txn.currency,
            f"{txn.card.card_name} ****{txn.card.card_last_four}" if txn.card else 'Cash',
            txn.expense_type,
            txn.merchant_group.name if txn.merchant_group else '',
            txn.source,
        ])

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="transactions.xlsx"'
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def transactions_import_excel(request):
    """Import transactions from .xlsx file"""
    import openpyxl
    from .models import Card as CardModel

    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception:
        return Response({'error': 'Invalid Excel file'}, status=status.HTTP_400_BAD_REQUEST)

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return Response({'error': 'Empty file'}, status=status.HTTP_400_BAD_REQUEST)

    # Normalize headers
    raw_headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    HEADER_MAP = {
        'date': 'date', 'transaction date': 'date', 'تاريخ': 'date',
        'type': 'type', 'transaction type': 'type', 'النوع': 'type',
        'merchant': 'merchant', 'merchant name': 'merchant', 'التاجر': 'merchant',
        'description': 'description', 'الوصف': 'description',
        'category': 'category', 'الفئة': 'category',
        'amount': 'amount', 'المبلغ': 'amount',
        'currency': 'currency', 'العملة': 'currency',
        'card': 'card', 'البطاقة': 'card',
        'expense type': 'expense_type', 'نوع المصروف': 'expense_type',
    }
    col_map = {}
    for i, h in enumerate(raw_headers):
        mapped = HEADER_MAP.get(h)
        if mapped and mapped not in col_map:
            col_map[mapped] = i

    if 'amount' not in col_map:
        return Response({'error': 'Missing required column: amount'}, status=status.HTTP_400_BAD_REQUEST)

    cards = {c.card_last_four: c for c in CardModel.objects.filter(user=request.user)}
    created, errors = 0, []

    from django.utils.dateparse import parse_datetime, parse_date
    from django.utils import timezone as tz
    from decimal import Decimal, InvalidOperation

    for row_idx, row in enumerate(rows[1:], 2):
        def get(key):
            idx = col_map.get(key)
            return row[idx] if idx is not None and idx < len(row) else None

        raw_amount = get('amount')
        try:
            amount = Decimal(str(raw_amount).replace(',', ''))
            if amount <= 0:
                raise ValueError
        except (InvalidOperation, ValueError, TypeError):
            errors.append(f'Row {row_idx}: invalid amount "{raw_amount}"')
            continue

        from datetime import datetime as _dt, time as _time, date as _date
        raw_date = get('date')
        txn_date = None
        if raw_date:
            if isinstance(raw_date, _dt):
                txn_date = tz.make_aware(raw_date) if tz.is_naive(raw_date) else raw_date
            elif isinstance(raw_date, _date):
                txn_date = tz.make_aware(_dt.combine(raw_date, _time.min))
            else:
                parsed = parse_datetime(str(raw_date)) or parse_date(str(raw_date))
                if parsed:
                    if isinstance(parsed, _dt):
                        txn_date = tz.make_aware(parsed) if tz.is_naive(parsed) else parsed
                    else:
                        txn_date = tz.make_aware(_dt.combine(parsed, _time.min))
        if not txn_date:
            txn_date = timezone.now()

        # Map card by last four digits
        card = None
        raw_card = str(get('card') or '')
        for last4, c in cards.items():
            if last4 in raw_card:
                card = c
                break

        raw_type = str(get('type') or 'PURCHASE').strip()
        valid_types = {t[0] for t in Transaction.TRANSACTION_TYPES}
        # try exact match first, then case-insensitive
        if raw_type in valid_types:
            txn_type = raw_type
        else:
            matched = next((v for v in valid_types if v.lower() == raw_type.lower()), None)
            txn_type = matched if matched else 'PURCHASE'

        from .models import Transaction as TxnModel
        TxnModel.objects.create(
            user=request.user,
            card=card,
            amount=amount,
            currency=str(get('currency') or 'AED').upper()[:3],
            transaction_type=txn_type,
            merchant_name=str(get('merchant') or '')[:255] or None,
            description=str(get('description') or '') or None,
            category=str(get('category') or '') or None,
            transaction_date=txn_date,
            source='excel_import',
        )
        created += 1

    return Response({'created': created, 'errors': errors, 'total_rows': len(rows) - 1})


# ─── Merchant Groups (Baskets) ────────────────────────────────────────────────

class MerchantGroupViewSet(viewsets.ModelViewSet):
    """CRUD for merchant groups (baskets) + nested rule management."""
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        from .models import MerchantGroup
        return MerchantGroup.objects.filter(user=self.request.user).prefetch_related('rules')

    def get_serializer_class(self):
        from .serializers import MerchantGroupSerializer
        return MerchantGroupSerializer

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=True, methods=['post'], url_path='rules')
    def add_rule(self, request, pk=None):
        """Add a merchant rule to this group."""
        from .models import MerchantGroup, MerchantRule
        from .serializers import MerchantRuleSerializer
        group = self.get_object()
        serializer = MerchantRuleSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(group=group)
            _classify_user_transactions(request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['delete'], url_path='rules/(?P<rule_id>[^/.]+)')
    def remove_rule(self, request, pk=None, rule_id=None):
        """Remove a merchant rule from this group."""
        from .models import MerchantRule
        group = self.get_object()
        try:
            rule = group.rules.get(id=rule_id)
        except MerchantRule.DoesNotExist:
            return Response({'detail': 'Rule not found.'}, status=status.HTTP_404_NOT_FOUND)
        rule.delete()
        _classify_user_transactions(request.user)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'], url_path='classify')
    def classify_all(self, request):
        """Re-run auto-classification on ALL user transactions."""
        updated = _classify_user_transactions(request.user)
        return Response({'classified': updated})

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        """Spending summary grouped by basket + expense_type breakdown."""
        from django.db.models import Sum, Count, Q
        from .models import Transaction
        from django.utils import timezone

        now = timezone.now()
        txns = Transaction.objects.filter(
            user=request.user, is_deleted=False,
            transaction_date__year=now.year,
            transaction_date__month=now.month,
        )
        company_total = txns.filter(expense_type='company').aggregate(t=Sum('amount'))['t'] or 0
        personal_total = txns.filter(expense_type='personal').aggregate(t=Sum('amount'))['t'] or 0
        unclassified_total = txns.filter(expense_type='unclassified').aggregate(t=Sum('amount'))['t'] or 0

        return Response({
            'month': now.strftime('%Y-%m'),
            'company': float(company_total),
            'personal': float(personal_total),
            'unclassified': float(unclassified_total),
            'total': float(company_total + personal_total + unclassified_total),
        })


def _classify_user_transactions(user):
    """
    Auto-classify transactions based on MerchantRule patterns.
    Returns the number of transactions updated.
    """
    from .models import MerchantRule, Transaction
    rules = MerchantRule.objects.filter(group__user=user).select_related('group')
    transactions = Transaction.objects.filter(user=user, is_deleted=False)
    updated = 0

    for txn in transactions:
        if not txn.merchant_name:
            continue
        merchant_lower = txn.merchant_name.lower()
        matched_group = None
        for rule in rules:
            name = rule.merchant_name.lower()
            if rule.match_type == 'exact' and merchant_lower == name:
                matched_group = rule.group
                break
            elif rule.match_type == 'contains' and name in merchant_lower:
                matched_group = rule.group
                break
            elif rule.match_type == 'starts_with' and merchant_lower.startswith(name):
                matched_group = rule.group
                break

        new_group = matched_group
        new_expense_type = matched_group.group_type if matched_group else 'unclassified'
        # Normalize mixed → unclassified for expense_type field
        if new_expense_type == 'mixed':
            new_expense_type = 'unclassified'

        if txn.merchant_group != new_group or txn.expense_type != new_expense_type:
            txn.merchant_group = new_group
            txn.expense_type = new_expense_type
            txn.save(update_fields=['merchant_group', 'expense_type'])
            updated += 1

    return updated


# ─── Cardholders (supplementary card holders) ─────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def cardholders_list(request):
    """
    Return all supplementary/joint cards with spending analytics per cardholder.
    """
    from .models import Card, Transaction
    from .services import encryption_service
    from django.db.models import Sum, Count, Max
    from django.utils import timezone

    cards = Card.objects.filter(
        user=request.user,
        card_ownership__in=['supplementary', 'joint'],
    )

    now = timezone.now()
    result = []
    for card in cards:
        # Decrypt cardholder name
        try:
            name = encryption_service.decrypt(card.cardholder_name_encrypted) if card.cardholder_name_encrypted else None
        except Exception:
            name = None

        txns = Transaction.objects.filter(card=card, is_deleted=False)
        monthly_txns = txns.filter(transaction_date__year=now.year, transaction_date__month=now.month)

        total_spent = txns.aggregate(t=Sum('amount'))['t'] or 0
        monthly_spent = monthly_txns.aggregate(t=Sum('amount'))['t'] or 0
        txn_count = txns.count()
        last_activity = txns.aggregate(d=Max('transaction_date'))['d']

        company_spent = txns.filter(expense_type='company').aggregate(t=Sum('amount'))['t'] or 0
        personal_spent = txns.filter(expense_type='personal').aggregate(t=Sum('amount'))['t'] or 0

        result.append({
            'card_id': str(card.id),
            'card_name': card.card_name,
            'card_last_four': card.card_last_four,
            'bank_name': card.bank_name,
            'card_ownership': card.card_ownership,
            'cardholder_name': name,
            'color_hex': card.color_hex,
            'credit_limit': float(card.credit_limit) if card.credit_limit else None,
            'total_spent': float(total_spent),
            'monthly_spent': float(monthly_spent),
            'company_spent': float(company_spent),
            'personal_spent': float(personal_spent),
            'transaction_count': txn_count,
            'last_activity': last_activity.isoformat() if last_activity else None,
        })

    result.sort(key=lambda x: x['monthly_spent'], reverse=True)
    return Response(result)


# ── Projects ──────────────────────────────────────────────────────────────────

class ProjectViewSet(viewsets.ModelViewSet):
    serializer_class = ProjectSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Project.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        project = serializer.save(user=self.request.user)
        audit_log(self.request, 'CREATE', 'Project', object_id=project.id, object_repr=project.name)

    def perform_update(self, serializer):
        project = serializer.save()
        audit_log(self.request, 'UPDATE', 'Project', object_id=project.id, object_repr=project.name)

    def destroy(self, request, pk=None):
        try:
            project = Project.objects.get(id=pk, user=request.user)
        except Project.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        project.is_deleted = True
        project.save(update_fields=['is_deleted'])
        audit_log(request, 'DELETE', 'Project', object_id=project.id, object_repr=project.name)
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Audit Log ─────────────────────────────────────────────────────────────────

class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = AuditLog.objects.filter(user=self.request.user).order_by('-created_at')
        model_name = self.request.query_params.get('model_name')
        if model_name:
            qs = qs.filter(model_name__iexact=model_name)
        action = self.request.query_params.get('action')
        if action:
            qs = qs.filter(action=action.upper())
        return qs[:500]

